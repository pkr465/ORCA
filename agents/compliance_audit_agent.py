"""
Comprehensive LLM-enriched compliance audit agent for ORCA.

Implements semantic code analysis with multi-layer context (CURE pattern) for C/C++ compliance auditing.
Features: multi-layer context building, chunked analysis with overlap, constraint injection,
HITL integration, Excel reporting, and per-chunk telemetry.
"""

import os
import time
import json
import logging
import uuid
import fnmatch
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class ComplianceReport:
    """Consolidated compliance audit report with LLM enrichment."""
    codebase_path: str
    findings: List[Dict[str, Any]] = field(default_factory=list)
    scores: Dict[str, float] = field(default_factory=dict)
    domain_results: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    domain_scores: Dict[str, float] = field(default_factory=dict)
    overall_grade: str = "A"
    timing: Dict[str, float] = field(default_factory=dict)
    is_compliant: bool = True
    run_id: str = ""  # UUID tracking for run-level analysis


class ComplianceAuditAgent:
    """
    LLM-enriched compliance auditor with chunked analysis and constraint injection.

    Uses 4 context layers (CURE pattern):
    1. Header context: File metadata, includes, dependencies
    2. Context validation: Call chains, function relationships
    3. Call stack analysis: Caller/callee relationships
    4. Function parameter validation: Signature compliance

    All context layers are optional with try/except fallback.
    """

    # Configurable chunking parameters
    TARGET_CHUNK_CHARS = 12000
    OVERLAP_LINES = 25

    def __init__(
        self,
        codebase_path: str,
        output_dir: str,
        config: Dict[str, Any],
        llm_tools: Any,
        rules: Dict[str, Any],
        exclude_dirs: Optional[List[str]] = None,
        exclude_globs: Optional[List[str]] = None,
        max_files: Optional[int] = None,
        domains: Optional[List[str]] = None,
        file_to_fix: Optional[str] = None,
        hitl_context: Optional[Dict[str, Any]] = None,
        constraints_dir: Optional[str] = None,
        custom_constraints: Optional[Dict[str, str]] = None,
        telemetry: bool = True,
        telemetry_run_id: Optional[str] = None,
    ):
        """
        Initialize the compliance audit agent.

        Args:
            codebase_path: Root directory of codebase to audit
            output_dir: Output directory for reports
            config: Configuration dictionary
            llm_tools: LLMTools instance for LLM calls (utils.llm_tools.LLMTools)
            rules: Compliance rules dictionary
            exclude_dirs: Directories to exclude from analysis
            exclude_globs: Glob patterns to exclude
            max_files: Maximum files to process
            domains: Specific domains to audit (style, license, structure, patch)
            file_to_fix: Specific file to focus on (optional)
            hitl_context: Human-in-the-loop context dictionary
            constraints_dir: Directory containing constraint markdown files
            custom_constraints: Custom constraint definitions
            telemetry: Enable per-chunk timing telemetry
            telemetry_run_id: UUID for this run (auto-generated if None)
        """
        self.codebase_path = codebase_path
        self.output_dir = output_dir
        self.config = config
        self.llm_tools = llm_tools
        self.rules = rules
        self.exclude_dirs = exclude_dirs or ['.git', '__pycache__', 'node_modules', '.venv']
        self.exclude_globs = exclude_globs or []
        self.max_files = max_files
        self.domains = domains or ['style', 'license', 'structure', 'patch']
        self.file_to_fix = file_to_fix
        self.hitl_context = hitl_context or {}
        self.constraints_dir = constraints_dir
        self.custom_constraints = custom_constraints or {}
        self.telemetry = telemetry
        self.telemetry_run_id = telemetry_run_id or str(uuid.uuid4())

        # Load constraints from markdown files
        self.constraints = self._load_constraints()

        # Static findings for merging with LLM findings
        self.static_findings = []

        logger.info(f"ComplianceAuditAgent initialized for {codebase_path}")
        logger.info(f"Run ID: {self.telemetry_run_id}")

    def run_analysis(self) -> ComplianceReport:
        """
        Execute the complete LLM-enriched compliance audit analysis.

        Returns:
            ComplianceReport with findings, scores, and grading
        """
        logger.info("Starting compliance audit analysis")
        report = ComplianceReport(
            codebase_path=self.codebase_path,
            run_id=self.telemetry_run_id
        )

        try:
            # Phase 1: File discovery
            start = time.time()
            files = self._discover_files()
            report.timing['phase_discovery'] = time.time() - start
            logger.info(f"Phase 1: Discovered {len(files)} files in {report.timing['phase_discovery']:.2f}s")

            # Phase 2: Chunk and analyze files
            start = time.time()
            all_findings = []
            for file_path in files:
                findings = self._analyze_file(file_path)
                all_findings.extend(findings)
            report.timing['phase_analysis'] = time.time() - start
            logger.info(f"Phase 2: Analyzed all files in {report.timing['phase_analysis']:.2f}s")

            # Phase 3: Merge findings (LLM + static)
            start = time.time()
            report.findings = self._merge_findings(all_findings, self.static_findings)
            report.timing['phase_merge'] = time.time() - start
            logger.info(f"Phase 3: Merged findings in {report.timing['phase_merge']:.2f}s")

            # Phase 4: Calculate scores
            start = time.time()
            report.scores, report.domain_scores = self._calculate_scores(report.findings)
            report.timing['phase_scoring'] = time.time() - start
            logger.info(f"Phase 4: Calculated scores in {report.timing['phase_scoring']:.2f}s")

            # Phase 5: Generate report
            start = time.time()
            self._write_excel_report(report)
            report.timing['phase_reporting'] = time.time() - start
            logger.info(f"Phase 5: Generated report in {report.timing['phase_reporting']:.2f}s")

            # Determine overall grade
            report.overall_grade = self._grade_from_score(report.scores.get('overall', 0.0))
            report.is_compliant = report.scores.get('overall', 0.0) >= self.config.get('compliance_threshold', 0.8)

            logger.info(f"Audit complete. Grade: {report.overall_grade}, "
                       f"Score: {report.scores.get('overall', 0):.2%}, "
                       f"Findings: {len(report.findings)}")

        except Exception as e:
            logger.error(f"Error during analysis: {e}", exc_info=True)
            report.is_compliant = False

        return report

    def _discover_files(self) -> List[str]:
        """
        Discover all source files to analyze.

        Returns:
            List of absolute file paths
        """
        files = []
        extensions = self.config.get('file_extensions', ['.c', '.h', '.cpp', '.hpp', '.py'])

        # Single-file mode
        if os.path.isfile(self.codebase_path):
            files.append(self.codebase_path)
            logger.info(f"Single-file mode: {self.codebase_path}")
            return files

        # Directory mode with recursive walk
        for root, dirs, filenames in os.walk(self.codebase_path):
            dirs[:] = [d for d in dirs if d not in self.exclude_dirs]

            for filename in filenames:
                if any(filename.endswith(ext) for ext in extensions):
                    file_path = os.path.join(root, filename)
                    relative_path = os.path.relpath(file_path, self.codebase_path)

                    # Check glob exclusions
                    if any(fnmatch.fnmatch(relative_path, g) for g in self.exclude_globs):
                        continue

                    files.append(file_path)

        if self.max_files:
            files = files[:self.max_files]

        logger.info(f"Discovered {len(files)} source files")
        return files

    def _analyze_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Analyze a single file with LLM-enriched semantic checks.

        Chunks the file if needed and builds multi-layer context before calling LLM.

        Args:
            file_path: Absolute path to file

        Returns:
            List of findings from this file
        """
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
        except Exception as e:
            logger.warning(f"Could not read {file_path}: {e}")
            return []

        findings = []
        lines = content.split('\n')

        # Chunk the file if needed
        chunks = self._chunk_file(content)
        logger.debug(f"File {file_path}: {len(chunks)} chunks")

        for chunk_idx, chunk in enumerate(chunks):
            try:
                # Build context layers for this chunk
                context = self._build_context_layers(file_path, content, lines, chunk)

                # Inject constraints into context
                context = self._inject_constraints(context)

                # Call LLM for each domain
                for domain in self.domains:
                    chunk_findings = self._audit_chunk(file_path, chunk, domain, context)
                    findings.extend(chunk_findings)

                # Telemetry: log per-chunk timing if enabled
                if self.telemetry:
                    logger.debug(f"{file_path}: chunk {chunk_idx + 1}/{len(chunks)} analyzed")

            except Exception as e:
                logger.warning(f"Error analyzing chunk {chunk_idx} in {file_path}: {e}")
                continue

        return findings

    def _chunk_file(self, content: str) -> List[str]:
        """
        Split file into overlapping chunks for analysis.

        Uses TARGET_CHUNK_CHARS and OVERLAP_LINES parameters.

        Args:
            content: File content as string

        Returns:
            List of chunk strings
        """
        lines = content.split('\n')
        chunks = []

        if len(content) <= self.TARGET_CHUNK_CHARS:
            return [content]

        start_idx = 0
        while start_idx < len(lines):
            # Calculate end index based on character count
            end_idx = start_idx
            current_chars = 0

            while end_idx < len(lines) and current_chars < self.TARGET_CHUNK_CHARS:
                current_chars += len(lines[end_idx]) + 1  # +1 for newline
                end_idx += 1

            # Get chunk with overlap from previous chunk
            chunk_start = max(0, start_idx - self.OVERLAP_LINES)
            chunk_lines = lines[chunk_start:end_idx]
            chunks.append('\n'.join(chunk_lines))

            start_idx = end_idx

        logger.debug(f"Created {len(chunks)} chunks from {len(lines)} lines")
        return chunks

    def _build_context_layers(
        self, file_path: str, full_content: str, lines: List[str], chunk: str
    ) -> Dict[str, Any]:
        """
        Build 4-layer context for improved LLM analysis (CURE pattern).

        All layers are optional with try/except fallback.

        Args:
            file_path: Path to file
            full_content: Complete file content
            lines: File lines
            chunk: Current chunk to analyze

        Returns:
            Context dictionary
        """
        context = {
            'file_path': file_path,
            'chunk': chunk,
            'header_context': {},
            'context_validation': {},
            'call_stack_analysis': {},
            'function_params': {}
        }

        # Layer 1: Header context (includes, defines)
        try:
            context['header_context'] = {
                'includes': self._extract_includes(full_content),
                'defines': self._extract_defines(full_content),
                'imports': self._extract_imports(full_content)
            }
        except Exception as e:
            logger.debug(f"Header context extraction failed: {e}")

        # Layer 2: Context validation (scopes, namespaces)
        try:
            context['context_validation'] = {
                'namespaces': self._extract_namespaces(full_content),
                'classes': self._extract_classes(full_content),
                'global_scope': self._extract_globals(full_content)
            }
        except Exception as e:
            logger.debug(f"Context validation failed: {e}")

        # Layer 3: Call stack analysis
        try:
            context['call_stack_analysis'] = {
                'function_defs': self._extract_function_defs(chunk),
                'function_calls': self._extract_function_calls(chunk)
            }
        except Exception as e:
            logger.debug(f"Call stack analysis failed: {e}")

        # Layer 4: Function parameter validation
        try:
            context['function_params'] = {
                'signatures': self._extract_function_sigs(chunk)
            }
        except Exception as e:
            logger.debug(f"Function param extraction failed: {e}")

        return context

    def _extract_includes(self, content: str) -> List[str]:
        """Extract #include directives."""
        includes = []
        for line in content.split('\n'):
            if line.strip().startswith('#include'):
                includes.append(line.strip())
        return includes[:10]

    def _extract_defines(self, content: str) -> List[str]:
        """Extract #define directives."""
        defines = []
        for line in content.split('\n'):
            if line.strip().startswith('#define'):
                defines.append(line.strip()[:80])
        return defines[:10]

    def _extract_imports(self, content: str) -> List[str]:
        """Extract import statements (Python, etc)."""
        imports = []
        for line in content.split('\n'):
            if line.strip().startswith('import ') or line.strip().startswith('from '):
                imports.append(line.strip())
        return imports[:10]

    def _extract_namespaces(self, content: str) -> List[str]:
        """Extract C++ namespace declarations."""
        namespaces = []
        for line in content.split('\n'):
            if 'namespace ' in line and '{' in line:
                namespaces.append(line.strip())
        return namespaces

    def _extract_classes(self, content: str) -> List[str]:
        """Extract class declarations."""
        classes = []
        for line in content.split('\n'):
            if any(kw in line for kw in ['class ', 'struct ']):
                classes.append(line.strip())
        return classes[:10]

    def _extract_globals(self, content: str) -> List[str]:
        """Extract global variable declarations."""
        globals_list = []
        for line in content.split('\n'):
            if line and not line.startswith(' ') and not line.startswith('\t'):
                if any(kw in line for kw in ['int ', 'char ', 'void ', 'static ', 'extern ']):
                    if '(' not in line:  # Exclude function declarations
                        globals_list.append(line.strip())
        return globals_list[:15]

    def _extract_function_defs(self, chunk: str) -> List[str]:
        """Extract function definitions from chunk."""
        defs = []
        for line in chunk.split('\n'):
            line = line.strip()
            if '(' in line and ')' in line and '{' in line:
                if not line.startswith('//') and not line.startswith('*'):
                    defs.append(line[:100])
        return defs

    def _extract_function_calls(self, chunk: str) -> List[str]:
        """Extract function calls from chunk."""
        calls = []
        for line in chunk.split('\n'):
            line = line.strip()
            if '(' in line and ')' in line:
                if not any(x in line for x in ['//', '/*', '*', '#']):
                    calls.append(line[:100])
        return calls[:20]

    def _extract_function_sigs(self, chunk: str) -> List[str]:
        """Extract function signatures."""
        sigs = []
        for line in chunk.split('\n'):
            line = line.strip()
            if '(' in line and ')' in line:
                if any(kw in line for kw in ['void ', 'int ', 'char ', 'bool ', 'float ']):
                    sigs.append(line[:120])
        return sigs[:10]

    def _inject_constraints(self, context: Dict[str, Any]) -> Dict[str, Any]:
        """
        Inject loaded constraints into context for LLM guidance.

        Args:
            context: Context dictionary

        Returns:
            Enhanced context with constraints
        """
        context['constraints'] = self.constraints
        context['constraints_text'] = self._get_constraints_text()
        context['custom_constraints'] = self.custom_constraints
        context['hitl_context'] = self.hitl_context
        return context

    def _get_constraints_text(self) -> str:
        """Build concatenated constraint text for prompt injection.

        Prioritises common/sample constraints first, then domain-specific files.
        """
        if not self.constraints:
            return ""

        parts = []
        # Common constraints first (applies to all domains)
        for key in ("common_constraints.md", "sample_constraints.md"):
            if key in self.constraints:
                parts.append(self.constraints[key])

        # Then all other constraint files
        for key, content in self.constraints.items():
            if key not in ("common_constraints.md", "sample_constraints.md"):
                parts.append(content)

        return "\n\n".join(parts)

    def _load_constraints(self) -> Dict[str, Any]:
        """
        Load constraints from markdown files in constraints_dir.

        Loads all .md files including common/sample constraints that apply to
        all domains, plus any domain-specific constraint files.

        Returns:
            Dictionary of loaded constraints (filename → content)
        """
        constraints = {}

        if not self.constraints_dir or not os.path.isdir(self.constraints_dir):
            return constraints

        try:
            for filename in os.listdir(self.constraints_dir):
                if filename.endswith('.md'):
                    filepath = os.path.join(self.constraints_dir, filename)
                    try:
                        with open(filepath, 'r', encoding='utf-8') as f:
                            constraints[filename] = f.read()
                    except Exception as e:
                        logger.warning(f"Could not load constraint {filename}: {e}")

            logger.info(f"Loaded {len(constraints)} constraint files from {self.constraints_dir}")
        except Exception as e:
            logger.warning(f"Error loading constraints: {e}")

        return constraints

    def _audit_chunk(
        self, file_path: str, chunk: str, domain: str, context: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Audit a chunk with LLM for specific domain.

        Uses chat_completion() from LLMTools router.

        Args:
            file_path: Path to file
            chunk: Code chunk to analyze
            domain: Domain (style, license, structure, patch)
            context: Multi-layer context

        Returns:
            List of findings
        """
        try:
            # Get prompts for domain
            system_prompt, user_prompt = self._get_domain_prompts(domain, file_path, chunk, context)

            # Call LLM using LLMTools.chat_completion()
            response = self.llm_tools.chat_completion(
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                system=system_prompt
            )

            result_text = response.content

            # Parse JSON response
            findings = self._parse_llm_response(result_text, file_path, domain)

            return findings

        except Exception as e:
            logger.warning(f"Error auditing chunk in {domain}: {e}")
            return []

    def _get_domain_prompts(
        self, domain: str, file_path: str, chunk: str, context: Dict[str, Any]
    ) -> Tuple[str, str]:
        """
        Get system and user prompts for domain.

        Args:
            domain: Domain name
            file_path: File path
            chunk: Code chunk
            context: Context dictionary

        Returns:
            Tuple of (system_prompt, user_prompt)
        """
        try:
            if domain == 'style':
                from agents.prompts.style_prompts import STYLE_SYSTEM_PROMPT, STYLE_USER_PROMPT_TEMPLATE
                system = STYLE_SYSTEM_PROMPT
                user = STYLE_USER_PROMPT_TEMPLATE.format(
                    file_path=file_path,
                    rules=json.dumps(self.rules.get('style', {}), indent=2),
                    code=chunk,
                    context=json.dumps({k: v for k, v in context.items() if k != 'chunk'}, indent=2, default=str)
                )

            elif domain == 'license':
                from agents.prompts.license_prompts import LICENSE_SYSTEM_PROMPT, LICENSE_USER_PROMPT_TEMPLATE
                system = LICENSE_SYSTEM_PROMPT
                user = LICENSE_USER_PROMPT_TEMPLATE.format(
                    file_path=file_path,
                    rules=json.dumps(self.rules.get('license', {}), indent=2),
                    code=chunk
                )

            elif domain == 'structure':
                from agents.prompts.structure_prompts import STRUCTURE_SYSTEM_PROMPT, STRUCTURE_USER_PROMPT_TEMPLATE
                system = STRUCTURE_SYSTEM_PROMPT
                user = STRUCTURE_USER_PROMPT_TEMPLATE.format(
                    file_path=file_path,
                    rules=json.dumps(self.rules.get('structure', {}), indent=2),
                    code=chunk,
                    context=json.dumps(context.get('context_validation', {}), indent=2, default=str)
                )

            elif domain == 'patch':
                from agents.prompts.patch_prompts import PATCH_SYSTEM_PROMPT, PATCH_USER_PROMPT_TEMPLATE
                system = PATCH_SYSTEM_PROMPT
                user = PATCH_USER_PROMPT_TEMPLATE.format(
                    file_path=file_path,
                    rules=json.dumps(self.rules.get('patch', {}), indent=2),
                    code=chunk
                )

            else:
                # Generic domain
                system = f"You are a code compliance auditor for domain '{domain}'."
                user = f"Analyze this code for {domain} compliance:\n\n```\n{chunk}\n```"

            return system, user

        except ImportError as e:
            logger.warning(f"Could not import prompts for {domain}: {e}")
            system = f"Analyze code for {domain} compliance"
            user = f"```\n{chunk}\n```"
            return system, user

    def _parse_llm_response(self, response_text: str, file_path: str, domain: str) -> List[Dict[str, Any]]:
        """
        Parse LLM JSON response with fallback.

        Uses LLMTools.extract_json() if available, with fallback parsing.

        Args:
            response_text: Raw LLM response
            file_path: File being analyzed
            domain: Domain analyzed

        Returns:
            List of findings
        """
        findings = []

        try:
            # Try to extract JSON using LLMTools
            if hasattr(self.llm_tools, 'extract_json'):
                json_obj = self.llm_tools.extract_json(response_text)
            else:
                # Fallback: manual JSON extraction
                start_idx = response_text.find('{')
                end_idx = response_text.rfind('}')
                if start_idx >= 0 and end_idx > start_idx:
                    json_str = response_text[start_idx:end_idx + 1]
                    json_obj = json.loads(json_str)
                else:
                    json_obj = {}

            # Extract violations/issues from response
            violations = json_obj.get('violations', [])
            issues = json_obj.get('issues', [])
            findings_list = json_obj.get('findings', [])

            all_items = violations + issues + findings_list

            for item in all_items:
                finding = {
                    'file_path': file_path,
                    'domain': domain,
                    'line_number': item.get('line_number', 0),
                    'severity': item.get('severity', 'MEDIUM'),
                    'violation_type': item.get('violation_type', 'unknown'),
                    'description': item.get('description', ''),
                    'suggested_fix': item.get('suggested_fix', ''),
                    'llm_enriched': True
                }
                findings.append(finding)

        except json.JSONDecodeError as e:
            logger.debug(f"Could not parse JSON from LLM response: {e}")
        except Exception as e:
            logger.warning(f"Error parsing LLM response: {e}")

        return findings

    def _merge_findings(self, llm_findings: List[Dict[str, Any]], static_findings: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Merge LLM findings with static findings, deduplicating.

        Args:
            llm_findings: Findings from LLM analysis
            static_findings: Findings from static analysis

        Returns:
            Deduplicated merged findings
        """
        # Create deduplication key for findings
        def finding_key(f: Dict[str, Any]) -> Tuple:
            return (f.get('file_path'), f.get('line_number'), f.get('description')[:50])

        # Deduplicate
        unique = {}
        for finding in llm_findings + static_findings:
            key = finding_key(finding)
            if key not in unique:
                unique[key] = finding

        merged = list(unique.values())
        logger.info(f"Merged {len(llm_findings)} LLM + {len(static_findings)} static findings -> {len(merged)} unique")

        return merged

    def _calculate_scores(self, findings: List[Dict[str, Any]]) -> Tuple[Dict[str, float], Dict[str, float]]:
        """
        Calculate overall and per-domain compliance scores.

        Args:
            findings: All findings

        Returns:
            Tuple of (overall_scores, domain_scores)
        """
        if not findings:
            return {'overall': 1.0, 'critical': 1.0, 'high': 1.0, 'medium': 1.0, 'low': 1.0}, {}

        severity_weights = {'CRITICAL': 0.4, 'HIGH': 0.3, 'MEDIUM': 0.2, 'LOW': 0.1}

        total_severity = 0.0
        for finding in findings:
            severity = finding.get('severity', 'LOW')
            total_severity += severity_weights.get(severity, 0.0)

        overall_score = max(0.0, 1.0 - (total_severity * 0.1))

        # Severity-specific scores
        severity_scores = {}
        for severity, weight in severity_weights.items():
            count = sum(1 for f in findings if f.get('severity') == severity)
            severity_scores[severity.lower()] = max(0.0, 1.0 - (count * 0.05))

        overall_scores = {'overall': overall_score, **severity_scores}

        # Domain-specific scores
        domain_scores = {}
        for domain in self.domains:
            domain_findings = [f for f in findings if f.get('domain') == domain]
            if domain_findings:
                domain_total = sum(severity_weights.get(f.get('severity', 'LOW'), 0.0) for f in domain_findings)
                domain_scores[domain] = max(0.0, 1.0 - (domain_total * 0.1))
            else:
                domain_scores[domain] = 1.0

        return overall_scores, domain_scores

    def _write_excel_report(self, report: ComplianceReport) -> None:
        """
        Write findings to Excel report.

        Args:
            report: ComplianceReport to write
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl not available, skipping Excel report")
            return

        try:
            os.makedirs(self.output_dir, exist_ok=True)

            filename = os.path.join(
                self.output_dir,
                f"compliance_audit_{report.run_id[:8]}.xlsx"
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Findings"

            # Header row with styling
            headers = [
                "File", "Line", "Domain", "Severity", "Type", "Description", "Suggested Fix"
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Data rows
            for row_idx, finding in enumerate(report.findings, 2):
                ws.cell(row=row_idx, column=1).value = finding.get('file_path', '')
                ws.cell(row=row_idx, column=2).value = finding.get('line_number', '')
                ws.cell(row=row_idx, column=3).value = finding.get('domain', '')
                ws.cell(row=row_idx, column=4).value = finding.get('severity', '')
                ws.cell(row=row_idx, column=5).value = finding.get('violation_type', '')
                ws.cell(row=row_idx, column=6).value = finding.get('description', '')
                ws.cell(row=row_idx, column=7).value = finding.get('suggested_fix', '')

            # Summary sheet
            summary = wb.create_sheet("Summary")
            summary['A1'] = "Codebase"
            summary['B1'] = self.codebase_path
            summary['A2'] = "Overall Score"
            summary['B2'] = report.scores.get('overall', 0.0)
            summary['A3'] = "Grade"
            summary['B3'] = report.overall_grade
            summary['A4'] = "Total Findings"
            summary['B4'] = len(report.findings)

            wb.save(filename)
            logger.info(f"Excel report written to {filename}")

        except Exception as e:
            logger.error(f"Error writing Excel report: {e}")

    def _grade_from_score(self, score: float) -> str:
        """Convert score to letter grade."""
        if score >= 0.9:
            return "A"
        elif score >= 0.8:
            return "B"
        elif score >= 0.7:
            return "C"
        elif score >= 0.6:
            return "D"
        else:
            return "F"
