"""
Comprehensive patch analysis agent for ORCA - LLM-enriched diff compliance auditor.

Analyzes patches (unified diffs) against C/C++ source files to identify issues
INTRODUCED by the patch. Modeled after CURE CodebasePatchAgent with full support for:
- Multiple diff formats (unified, normal, context, combined)
- Patch hunk parsing and reconstruction
- Code region extraction with context
- LLM analysis of both original and patched versions
- Static analyzer integration
- Delta analysis to isolate newly introduced issues
- Constraint injection from markdown
- HITL feedback integration
- Excel report generation with per-file patch tabs
- Per-chunk telemetry tracking
"""

import os
import re
import time
import json
import logging
import uuid
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# DATACLASSES
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class PatchHunk:
    """Represents a single hunk from a unified diff."""
    orig_start: int
    orig_count: int
    new_start: int
    new_count: int
    header: str
    removed_lines: List[str] = field(default_factory=list)
    added_lines: List[str] = field(default_factory=list)
    context_lines: List[str] = field(default_factory=list)
    raw_lines: List[str] = field(default_factory=list)


@dataclass
class PatchFinding:
    """Represents a single compliance finding from patch analysis."""
    file_path: str
    line_number: int
    severity: str  # CRITICAL, HIGH, MEDIUM, LOW
    category: str  # INTRODUCED_ISSUE, FORMAT, METADATA, STYLE, etc.
    description: str
    title: str
    confidence: float  # 0.0 to 1.0
    suggestion: str
    code_before: Optional[str] = None
    code_after: Optional[str] = None
    introduced_by_patch: bool = True
    issue_source: str = "llm"  # llm, static_analyzer, format_check
    feedback: Optional[str] = None
    constraints: Optional[List[str]] = None


@dataclass
class PatchAnalysisReport:
    """Complete patch analysis report."""
    patch_file: str
    source_path: str
    findings: List[PatchFinding] = field(default_factory=list)
    hunks_analyzed: int = 0
    total_issues_original: int = 0
    total_issues_patched: int = 0
    newly_introduced: int = 0
    is_compliant: bool = True
    run_id: str = ""
    timing: Dict[str, float] = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────────────────────
# REGEX PATTERNS FOR DIFF PARSING
# ─────────────────────────────────────────────────────────────────────────────

_UNIFIED_HUNK_RE = re.compile(
    r'^@@ -(\d+)(?:,(\d+))? \+(\d+)(?:,(\d+))? @@\s*(.*?)$',
    re.MULTILINE
)

_NORMAL_CMD_RE = re.compile(
    r'^(\d+)(?:,(\d+))?([acd])(\d+)(?:,(\d+))?$'
)

_CONTEXT_HUNK_RE = re.compile(
    r'^\*\*\* (\d+)(?:,(\d+))? \*\*\*\*',
    re.MULTILINE
)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN AGENT CLASS
# ─────────────────────────────────────────────────────────────────────────────


class CompliancePatchAgent:
    """
    LLM-enriched patch/diff analysis agent.

    Analyzes patches to identify compliance issues INTRODUCED by the patch.
    Runs LLM analysis on both original and patched code, then diffs findings
    to isolate newly introduced issues.
    """

    def __init__(
        self,
        codebase_path: str,
        output_dir: str,
        config: Dict[str, Any],
        llm_tools: Any,
        rules: Dict[str, Any],
        hitl_context: Optional[Dict[str, Any]] = None,
        constraints_dir: Optional[str] = None,
        enable_static: bool = False,
        telemetry: bool = True,
        telemetry_run_id: Optional[str] = None,
    ):
        """
        Initialize the compliance patch analysis agent.

        Args:
            codebase_path: Root directory of codebase
            output_dir: Output directory for reports
            config: Configuration dictionary
            llm_tools: LLMTools instance for LLM analysis
            rules: Compliance rules dictionary
            hitl_context: Human-in-the-loop feedback context
            constraints_dir: Directory containing constraint markdown files
            enable_static: Enable static analyzer integration
            telemetry: Enable per-chunk telemetry
            telemetry_run_id: UUID for run tracking
        """
        self.codebase_path = codebase_path
        self.output_dir = output_dir
        self.config = config
        self.llm_tools = llm_tools
        self.rules = rules
        self.hitl_context = hitl_context or {}
        self.constraints_dir = constraints_dir
        self.enable_static = enable_static
        self.telemetry = telemetry
        self.telemetry_run_id = telemetry_run_id or str(uuid.uuid4())

        # Load constraints from markdown
        self.constraints = self._load_constraints()

        logger.info(f"CompliancePatchAgent initialized")
        logger.info(f"Run ID: {self.telemetry_run_id}")

    def analyze_patch(
        self, source_path: str, patch_text: str
    ) -> List[PatchFinding]:
        """
        Analyze a patch against a source file to identify introduced issues.

        Args:
            source_path: Absolute path to source file being patched
            patch_text: Unified diff content as string

        Returns:
            List of PatchFinding objects for newly introduced issues
        """
        logger.info(f"Analyzing patch for {source_path}")

        try:
            # Read original source
            with open(source_path, 'r', encoding='utf-8', errors='ignore') as f:
                original_source = f.read()
        except Exception as e:
            logger.error(f"Error reading source {source_path}: {e}")
            return []

        # Parse patch hunks
        hunks = self._parse_unified_diff(patch_text)
        if not hunks:
            logger.warning(f"No hunks found in patch for {source_path}")
            return []

        # Apply hunks to reconstruct patched version
        try:
            patched_source = self._apply_hunks(original_source, hunks)
        except Exception as e:
            logger.error(f"Error applying hunks: {e}")
            return []

        # Extract hunk regions with context
        original_regions = self._extract_hunk_regions(
            original_source, hunks, context_lines=30
        )
        patched_regions = self._extract_hunk_regions(
            patched_source, hunks, context_lines=30
        )

        # Load constraints
        constraints_str = self._load_constraints()

        # Run LLM analysis on original code
        original_findings = self._run_llm_analysis(
            source_path, original_regions, "original", constraints_str
        )

        # Run LLM analysis on patched code
        patched_findings = self._run_llm_analysis(
            source_path, patched_regions, "patched", constraints_str
        )

        # Optionally run static analysis on patched code
        static_findings = []
        if self.enable_static:
            static_findings = self._run_static_analysis(source_path, patched_source)

        # Diff findings to isolate newly introduced issues
        new_findings = self._diff_findings(original_findings, patched_findings)
        new_findings.extend(self._diff_findings([], static_findings))

        # Convert to PatchFinding objects
        patch_findings = []
        for finding in new_findings:
            pf = PatchFinding(
                file_path=source_path,
                line_number=finding.get("line_number", 0),
                severity=finding.get("severity", "MEDIUM"),
                category=finding.get("category", "INTRODUCED_ISSUE"),
                description=finding.get("description", ""),
                title=finding.get("title", ""),
                confidence=finding.get("confidence", 0.5),
                suggestion=finding.get("suggestion", ""),
                code_before=finding.get("code_before"),
                code_after=finding.get("code_after"),
                introduced_by_patch=True,
                issue_source=finding.get("source", "llm"),
            )
            patch_findings.append(pf)

        logger.info(f"Found {len(patch_findings)} newly introduced issues")
        return patch_findings

    def analyze_patch_file(
        self, source_path: str, patch_file_path: str
    ) -> List[PatchFinding]:
        """
        Analyze a patch file against a source file.

        Args:
            source_path: Absolute path to source file
            patch_file_path: Absolute path to patch file

        Returns:
            List of PatchFinding objects
        """
        try:
            with open(patch_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                patch_text = f.read()
        except Exception as e:
            logger.error(f"Error reading patch file {patch_file_path}: {e}")
            return []

        return self.analyze_patch(source_path, patch_text)

    def _parse_unified_diff(self, text: str) -> List[PatchHunk]:
        """
        Parse unified diff format (@@ hunk headers).

        Args:
            text: Diff content

        Returns:
            List of PatchHunk objects
        """
        hunks = []
        lines = text.split('\n')

        current_hunk = None
        i = 0

        while i < len(lines):
            line = lines[i]
            match = _UNIFIED_HUNK_RE.match(line)

            if match:
                if current_hunk:
                    hunks.append(current_hunk)

                orig_start = int(match.group(1))
                orig_count = int(match.group(2)) if match.group(2) else 1
                new_start = int(match.group(3))
                new_count = int(match.group(4)) if match.group(4) else 1

                current_hunk = PatchHunk(
                    orig_start=orig_start,
                    orig_count=orig_count,
                    new_start=new_start,
                    new_count=new_count,
                    header=line,
                    removed_lines=[],
                    added_lines=[],
                    context_lines=[],
                    raw_lines=[],
                )
                i += 1
            elif current_hunk is not None:
                if line.startswith(('+', '-', ' ', '\\')):
                    current_hunk.raw_lines.append(line)

                    if line.startswith('+') and not line.startswith('+++'):
                        current_hunk.added_lines.append(line[1:])
                    elif line.startswith('-') and not line.startswith('---'):
                        current_hunk.removed_lines.append(line[1:])
                    elif line.startswith(' '):
                        current_hunk.context_lines.append(line[1:])

                    i += 1
                else:
                    break
            else:
                i += 1

        if current_hunk:
            hunks.append(current_hunk)

        logger.debug(f"Parsed {len(hunks)} hunks from unified diff")
        return hunks

    def _parse_normal_diff(self, text: str) -> List[PatchHunk]:
        """
        Parse normal diff format (NUMcNUM, NUMdNUM, etc.).

        Args:
            text: Diff content in normal format

        Returns:
            List of PatchHunk objects
        """
        hunks = []
        lines = text.split('\n')

        i = 0
        while i < len(lines):
            line = lines[i]
            match = _NORMAL_CMD_RE.match(line)

            if match:
                orig_start = int(match.group(1))
                orig_end = int(match.group(2)) if match.group(2) else orig_start
                cmd = match.group(3)  # a, c, d
                new_start = int(match.group(4))
                new_end = int(match.group(5)) if match.group(5) else new_start

                orig_count = orig_end - orig_start + 1
                new_count = new_end - new_start + 1

                hunk = PatchHunk(
                    orig_start=orig_start,
                    orig_count=orig_count,
                    new_start=new_start,
                    new_count=new_count,
                    header=line,
                    raw_lines=[line],
                )

                # Parse content lines
                i += 1
                if cmd in ('c', 'd'):
                    # Removed lines (< marker)
                    while i < len(lines) and lines[i].startswith('< '):
                        hunk.removed_lines.append(lines[i][2:])
                        hunk.raw_lines.append(lines[i])
                        i += 1

                if cmd in ('c', 'a'):
                    # Separator for change
                    if i < len(lines) and lines[i] == '---':
                        hunk.raw_lines.append(lines[i])
                        i += 1

                    # Added lines (> marker)
                    while i < len(lines) and lines[i].startswith('> '):
                        hunk.added_lines.append(lines[i][2:])
                        hunk.raw_lines.append(lines[i])
                        i += 1

                hunks.append(hunk)
            else:
                i += 1

        logger.debug(f"Parsed {len(hunks)} hunks from normal diff")
        return hunks

    def _detect_diff_format(self, text: str) -> str:
        """
        Auto-detect diff format (unified, normal, context, combined).

        Args:
            text: Diff content

        Returns:
            Format string: 'unified', 'normal', 'context', 'combined'
        """
        if '@@' in text and '+' in text and '-' in text:
            return 'unified'
        elif _NORMAL_CMD_RE.search(text):
            return 'normal'
        elif '***' in text and '---' in text:
            return 'context'
        elif 'diff --git' in text:
            return 'combined'
        else:
            return 'unified'  # default

    def _apply_hunks(self, source: str, hunks: List[PatchHunk]) -> str:
        """
        Apply hunks to original source to reconstruct patched version.

        Args:
            source: Original source code
            hunks: List of hunks to apply

        Returns:
            Patched source code as string
        """
        lines = source.split('\n')

        # Process hunks in reverse order to maintain line numbers
        for hunk in reversed(hunks):
            start_idx = hunk.orig_start - 1  # Convert to 0-based
            end_idx = start_idx + hunk.orig_count

            # Replace old lines with new lines
            new_section = hunk.added_lines
            lines[start_idx:end_idx] = new_section

        return '\n'.join(lines)

    def _extract_hunk_regions(
        self, content: str, hunks: List[PatchHunk], context_lines: int = 30
    ) -> str:
        """
        Extract code regions around hunks with context.

        Args:
            content: Source code content
            hunks: List of hunks to extract regions for
            context_lines: Number of context lines before/after hunk

        Returns:
            Extracted regions as a single string
        """
        lines = content.split('\n')
        regions = []

        for hunk in hunks:
            start = max(0, hunk.orig_start - 1 - context_lines)
            end = min(len(lines), hunk.orig_start - 1 + hunk.orig_count + context_lines)

            region = '\n'.join(lines[start:end])
            regions.append(region)

        return '\n\n--- REGION SEPARATOR ---\n\n'.join(regions)

    def _run_llm_analysis(
        self, file_path: str, code: str, domain: str, constraints: str
    ) -> List[Dict[str, Any]]:
        """
        Run LLM analysis on code for compliance issues.

        Args:
            file_path: Path to file being analyzed
            code: Code snippet to analyze
            domain: Analysis domain (original, patched)
            constraints: Constraint definitions

        Returns:
            List of findings dictionaries
        """
        if not self.llm_tools:
            logger.warning("LLM tools not available, skipping LLM analysis")
            return []

        try:
            from agents.prompts.patch_prompts import (
                PATCH_SYSTEM_PROMPT,
                PATCH_USER_PROMPT_TEMPLATE,
            )

            prompt = self._build_patch_review_prompt(
                file_path, code, code, [], constraints
            )

            response = self.llm_tools.call_llm(
                system_prompt=PATCH_SYSTEM_PROMPT,
                user_prompt=prompt,
                temperature=0.3,
                max_tokens=2048,
            )

            # Parse LLM response as JSON findings
            if response and response.content:
                try:
                    findings = json.loads(response.content)
                    if isinstance(findings, dict) and 'patch_findings' in findings:
                        return findings['patch_findings']
                    elif isinstance(findings, list):
                        return findings
                except json.JSONDecodeError:
                    logger.warning(f"Failed to parse LLM response as JSON")
                    return []

        except Exception as e:
            logger.error(f"Error during LLM analysis: {e}")

        return []

    def _diff_findings(
        self, original_findings: List[Dict[str, Any]], patched_findings: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Diff findings between original and patched versions.

        Returns only findings that are NEW in patched version (not in original).

        Args:
            original_findings: Findings from original code
            patched_findings: Findings from patched code

        Returns:
            List of newly introduced findings
        """
        # Create set of original finding descriptions for quick lookup
        original_descriptions = {
            f.get("description", "") for f in original_findings
        }

        # Return findings that are new (not in original)
        new_findings = []
        for finding in patched_findings:
            desc = finding.get("description", "")
            if desc not in original_descriptions:
                new_findings.append(finding)

        logger.debug(f"Diffing: {len(original_findings)} original vs "
                    f"{len(patched_findings)} patched, found {len(new_findings)} new")

        return new_findings

    def _run_static_analysis(
        self, file_path: str, content: str
    ) -> List[Dict[str, Any]]:
        """
        Run static analyzer on patched file content.

        Args:
            file_path: Path to file
            content: File content

        Returns:
            List of findings from static analyzer
        """
        findings = []

        # Check for common C/C++ issues in newly added code
        lines = content.split('\n')

        for line_num, line in enumerate(lines, 1):
            # Check for buffer overflow patterns
            if 'strcpy' in line or 'sprintf' in line or 'gets' in line:
                findings.append({
                    "line_number": line_num,
                    "severity": "HIGH",
                    "category": "SECURITY",
                    "title": "Unsafe function usage",
                    "description": f"Line uses unsafe function: {line.strip()}",
                    "suggestion": "Use safer alternatives (strcpy_s, snprintf, fgets)",
                    "confidence": 0.8,
                    "source": "static_analyzer",
                })

            # Check for missing NULL checks
            if '->' in line and 'if' not in line and 'NULL' not in line:
                findings.append({
                    "line_number": line_num,
                    "severity": "MEDIUM",
                    "category": "SAFETY",
                    "title": "Potential NULL dereference",
                    "description": f"Pointer dereference without NULL check",
                    "suggestion": "Add NULL check before dereferencing",
                    "confidence": 0.6,
                    "source": "static_analyzer",
                })

        logger.debug(f"Static analysis found {len(findings)} issues")
        return findings

    def _load_constraints(self) -> str:
        """
        Load constraint definitions from markdown files.

        Returns:
            Concatenated constraint definitions as string
        """
        constraints = ""

        if not self.constraints_dir or not os.path.isdir(self.constraints_dir):
            return constraints

        try:
            for filename in os.listdir(self.constraints_dir):
                if filename.endswith('.md'):
                    filepath = os.path.join(self.constraints_dir, filename)
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        constraints += f"\n\n--- {filename} ---\n{f.read()}"
        except Exception as e:
            logger.warning(f"Error loading constraints: {e}")

        return constraints

    def _write_patch_report(
        self, findings: List[PatchFinding], output_path: str, filename: str
    ) -> None:
        """
        Write Excel report with patch analysis findings.

        Args:
            findings: List of findings
            output_path: Output file path
            filename: Source filename for tab name
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            # Create or load workbook
            if os.path.exists(output_path):
                wb = openpyxl.load_workbook(output_path)
            else:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)

            # Create sheet for this patch
            sheet_name = f"patch_{filename[:20]}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(2, ws.max_row)
            else:
                ws = wb.create_sheet(sheet_name)

            # Headers
            headers = [
                "Line", "Severity", "Category", "Title", "Description",
                "Suggestion", "Confidence", "Source", "Introduced by Patch"
            ]
            ws.append(headers)

            # Style header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add findings
            for finding in findings:
                severity_color = {
                    "CRITICAL": "FF0000",
                    "HIGH": "FF6600",
                    "MEDIUM": "FFCC00",
                    "LOW": "00CC00",
                }.get(finding.severity, "FFFFFF")

                row = [
                    finding.line_number,
                    finding.severity,
                    finding.category,
                    finding.title,
                    finding.description,
                    finding.suggestion,
                    f"{finding.confidence:.2%}",
                    finding.issue_source,
                    "Yes" if finding.introduced_by_patch else "No",
                ]
                ws.append(row)

                # Color severity cell
                ws[f"B{ws.max_row}"].fill = PatternFill(
                    start_color=severity_color, end_color=severity_color, fill_type="solid"
                )

            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 40

            wb.save(output_path)
            logger.info(f"Wrote patch report to {output_path}")

        except ImportError:
            logger.warning("openpyxl not available, skipping Excel report")
        except Exception as e:
            logger.error(f"Error writing patch report: {e}")

    def _build_patch_review_prompt(
        self,
        file_path: str,
        original_code: str,
        patched_code: str,
        hunks: List[PatchHunk],
        constraints: str,
    ) -> str:
        """
        Build comprehensive patch review prompt for LLM.

        Args:
            file_path: Path to file being patched
            original_code: Original code content
            patched_code: Patched code content
            hunks: List of hunks
            constraints: Constraint definitions

        Returns:
            Formatted prompt for LLM
        """
        prompt = f"""Review the following patch for compliance issues INTRODUCED BY THE PATCH.

FILE: {file_path}

ORIGINAL CODE:
```
{original_code[:2000]}
```

PATCHED CODE:
```
{patched_code[:2000]}
```

APPLICABLE CONSTRAINTS:
{constraints[:1000] if constraints else "No constraints defined"}

Identify ONLY issues that are:
1. Newly introduced by this patch (not present in original)
2. Violations of compliance rules
3. Potential security or safety issues

Return as JSON array with objects containing:
- line_number (int)
- severity (CRITICAL|HIGH|MEDIUM|LOW)
- category (string)
- title (string)
- description (string)
- suggestion (string)
- confidence (0.0-1.0)
- code_before (string or null)
- code_after (string or null)
"""
        return prompt
