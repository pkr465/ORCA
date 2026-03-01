"""Excel report generation adapter using openpyxl."""
import os
from datetime import datetime
from typing import Dict, List, Optional, Any
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from agents.adapters.base_adapter import BaseComplianceAdapter, AdapterResult

class ExcelTheme:
    """Professional Excel styling theme."""
    
    HEADER_FILL = PatternFill(start_color="000F3A", end_color="000F3A", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    SEVERITY_COLORS = {
        "critical": "FFC7CE",  # Red
        "high": "FFEB9C",      # Orange
        "medium": "FFF2CC",    # Yellow
        "low": "E2EFDA",       # Green
    }
    
    ALTERNATING_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def get_severity_fill(severity: str) -> PatternFill:
        """Get fill color for severity level."""
        color = ExcelTheme.SEVERITY_COLORS.get(severity, "FFFFFF")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

class ExcelReportAdapter(BaseComplianceAdapter):
    """Adapter for generating comprehensive Excel compliance reports."""
    
    def __init__(self, rules: dict, config: dict):
        """Initialize Excel report adapter.
        
        Args:
            rules: Configuration rules
            config: Global configuration
        """
        super().__init__(rules, config)
        if not HAS_OPENPYXL:
            self.logger.warning("openpyxl not installed, Excel reports unavailable")
    
    def analyze(self, file_cache: Dict[str, str], **kwargs) -> AdapterResult:
        """Excel adapter doesn't analyze, it reports.
        
        Args:
            file_cache: Dict of files (unused)
            **kwargs: Additional arguments (unused)
            
        Returns:
            AdapterResult indicating tool availability
        """
        return AdapterResult(
            score=100.0,
            grade="A",
            domain="reporting",
            findings=[],
            summary={"status": "report_generator"},
            tool_available=HAS_OPENPYXL,
            tool_name="excel_reporter"
        )
    
    def generate_report(self, compliance_report, output_path: str) -> bool:
        """Generate comprehensive multi-tab Excel report.

        Args:
            compliance_report: ComplianceReport dataclass or dictionary containing compliance analysis results with structure:
                {
                    "overall_score": float,
                    "overall_grade": str,
                    "summary": {...},
                    "findings": [...],
                    "domains": {
                        "code_style": {"score": float, "findings": [...]},
                        "license_compliance": {...},
                        ...
                    }
                }
            output_path: Path where Excel file should be written

        Returns:
            True if successful, False otherwise
        """
        if not HAS_OPENPYXL:
            self.logger.error("openpyxl not available, cannot generate Excel report")
            return False

        try:
            # Convert ComplianceReport dataclass to dict format
            report_dict = self._to_dict(compliance_report)

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create sheets
            self._create_summary_sheet(wb, report_dict)
            self._create_style_violations_sheet(wb, report_dict)
            self._create_license_violations_sheet(wb, report_dict)
            self._create_structure_violations_sheet(wb, report_dict)
            self._create_patch_violations_sheet(wb, report_dict)
            self._create_decision_trail_sheet(wb, report_dict)
            
            # Save
            wb.save(output_path)
            self.logger.info(f"Excel report generated: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            return False

    def _to_dict(self, obj) -> Dict[str, Any]:
        """Convert ComplianceReport dataclass to dict format.

        Args:
            obj: ComplianceReport dataclass or already a dict

        Returns:
            Dictionary representation
        """
        if isinstance(obj, dict):
            return obj

        # Convert dataclass to dict
        from dataclasses import asdict, is_dataclass
        if is_dataclass(obj) and not isinstance(obj, type):
            result = asdict(obj)
            # Convert Finding objects to dicts
            if 'findings' in result and result['findings']:
                result['findings'] = [
                    asdict(f) if is_dataclass(f) and not isinstance(f, type) else f
                    for f in result['findings']
                ]
            # Convert by_domain values
            if 'by_domain' in result and result['by_domain']:
                result['by_domain'] = {
                    k: [asdict(f) if is_dataclass(f) and not isinstance(f, type) else f for f in v]
                    for k, v in result['by_domain'].items()
                }
            return result
        return obj
    
    def _create_summary_sheet(self, wb: Any, report: Dict[str, Any]) -> None:
        """Create Summary sheet with overall metrics and chart data.
        
        Args:
            wb: Workbook object
            report: Compliance report
        """
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "COMPLIANCE REPORT SUMMARY"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Generated timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=9)
        
        # Overall metrics
        row = 4
        ws[f'A{row}'] = "Overall Grade:"
        ws[f'B{row}'] = report.get('overall_grade', 'N/A')
        ws[f'B{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f'B{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        row += 1
        ws[f'A{row}'] = "Overall Score:"
        ws[f'B{row}'] = report.get('overall_score', 0)
        ws[f'B{row}'].number_format = '0.0'
        
        row += 2
        ws[f'A{row}'] = "Per-Domain Scores:"
        ws[f'A{row}'].font = Font(bold=True)
        
        # Domain scores table
        row += 1
        headers = ["Domain", "Score", "Grade", "Files", "Issues"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = ExcelTheme.HEADER_FILL
            cell.font = ExcelTheme.HEADER_FONT
            cell.border = ExcelTheme.BORDER
        
        row += 1
        domains = report.get('domains', {})
        for domain_name, domain_data in domains.items():
            ws.cell(row=row, column=1).value = domain_name
            ws.cell(row=row, column=2).value = domain_data.get('score', 0)
            ws.cell(row=row, column=3).value = domain_data.get('grade', 'N/A')
            ws.cell(row=row, column=4).value = len(domain_data.get('findings', []))
            ws.cell(row=row, column=5).value = len(domain_data.get('findings', []))
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = ExcelTheme.BORDER
            row += 1
        
        # Finding distribution
        row += 2
        ws[f'A{row}'] = "Finding Distribution:"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        severity_counts = defaultdict(int)
        for finding in report.get('findings', []):
            severity = finding.get('severity', 'unknown')
            severity_counts[severity] += 1
        
        for severity in ['critical', 'high', 'medium', 'low']:
            ws[f'A{row}'] = severity.capitalize()
            ws[f'B{row}'] = severity_counts.get(severity, 0)
            ws[f'B{row}'].fill = ExcelTheme.get_severity_fill(severity)
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    def _create_style_violations_sheet(self, wb: Any, report: Dict[str, Any]) -> None:
        """Create style_violations sheet.
        
        Args:
            wb: Workbook object
            report: Compliance report
        """
        ws = wb.create_sheet("style_violations")
        
        headers = ["File", "Line", "Rule ID", "Severity", "Description", "Suggested Fix", "Action"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = ExcelTheme.HEADER_FILL
            cell.font = ExcelTheme.HEADER_FONT
            cell.border = ExcelTheme.BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        
        # Get style violations
        style_findings = []
        for domain_name, domain_data in report.get('domains', {}).items():
            if 'style' in domain_name.lower():
                style_findings.extend(domain_data.get('findings', []))
        
        row = 2
        for finding in style_findings:
            ws.cell(row=row, column=1).value = finding.get('file_path', '')
            ws.cell(row=row, column=2).value = finding.get('line_number', '')
            ws.cell(row=row, column=3).value = finding.get('rule_id', '')
            
            severity = finding.get('severity', 'low')
            ws.cell(row=row, column=4).value = severity
            ws.cell(row=row, column=4).fill = ExcelTheme.get_severity_fill(severity)
            
            ws.cell(row=row, column=5).value = finding.get('message', '')
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
            
            ws.cell(row=row, column=6).value = finding.get('suggestion', '')
            ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
            
            # Action column (dropdown would need data validation)
            ws.cell(row=row, column=7).value = ""
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = ExcelTheme.BORDER
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 15
        
        # Freeze header
        ws.freeze_panes = "A2"
    
    def _create_license_violations_sheet(self, wb: Any, report: Dict[str, Any]) -> None:
        """Create license_violations sheet.
        
        Args:
            wb: Workbook object
            report: Compliance report
        """
        ws = wb.create_sheet("license_violations")
        
        headers = ["File", "Line", "Issue", "Current Header", "Expected Header", "Severity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = ExcelTheme.HEADER_FILL
            cell.font = ExcelTheme.HEADER_FONT
            cell.border = ExcelTheme.BORDER
        
        # Get license findings
        license_findings = []
        domain_data = report.get('domains', {}).get('license_compliance', {})
        license_findings = domain_data.get('findings', [])
        
        row = 2
        for finding in license_findings:
            ws.cell(row=row, column=1).value = finding.get('file_path', '')
            ws.cell(row=row, column=2).value = finding.get('line_number', '')
            ws.cell(row=row, column=3).value = finding.get('message', '')
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=4).value = ""
            ws.cell(row=row, column=5).value = finding.get('suggestion', '')
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
            
            severity = finding.get('severity', 'low')
            ws.cell(row=row, column=6).value = severity
            ws.cell(row=row, column=6).fill = ExcelTheme.get_severity_fill(severity)
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = ExcelTheme.BORDER
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        
        ws.freeze_panes = "A2"
    
    def _create_structure_violations_sheet(self, wb: Any, report: Dict[str, Any]) -> None:
        """Create structure_violations sheet.
        
        Args:
            wb: Workbook object
            report: Compliance report
        """
        ws = wb.create_sheet("structure_violations")
        
        headers = ["File", "Line", "Rule ID", "Issue", "Module", "Suggested Fix", "Severity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = ExcelTheme.HEADER_FILL
            cell.font = ExcelTheme.HEADER_FONT
            cell.border = ExcelTheme.BORDER
        
        # Get structure findings
        structure_findings = []
        domain_data = report.get('domains', {}).get('code_structure', {})
        structure_findings = domain_data.get('findings', [])
        
        row = 2
        for finding in structure_findings:
            file_path = finding.get('file_path', '')
            module = file_path.split('/')[0] if '/' in file_path else "root"
            
            ws.cell(row=row, column=1).value = file_path
            ws.cell(row=row, column=2).value = finding.get('line_number', '')
            ws.cell(row=row, column=3).value = finding.get('rule_id', '')
            ws.cell(row=row, column=4).value = finding.get('message', '')
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=5).value = module
            ws.cell(row=row, column=6).value = finding.get('suggestion', '')
            ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
            
            severity = finding.get('severity', 'low')
            ws.cell(row=row, column=7).value = severity
            ws.cell(row=row, column=7).fill = ExcelTheme.get_severity_fill(severity)
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = ExcelTheme.BORDER
            
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        
        ws.freeze_panes = "A2"
    
    def _create_patch_violations_sheet(self, wb: Any, report: Dict[str, Any]) -> None:
        """Create patch_violations sheet.
        
        Args:
            wb: Workbook object
            report: Compliance report
        """
        ws = wb.create_sheet("patch_violations")
        
        headers = ["Commit", "Line", "Rule ID", "Issue", "Hunk", "Suggested Fix", "Severity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = ExcelTheme.HEADER_FILL
            cell.font = ExcelTheme.HEADER_FONT
            cell.border = ExcelTheme.BORDER
        
        # Get patch findings
        patch_findings = []
        domain_data = report.get('domains', {}).get('patch_format', {})
        patch_findings = domain_data.get('findings', [])
        
        row = 2
        for finding in patch_findings:
            ws.cell(row=row, column=1).value = finding.get('file_path', '')
            ws.cell(row=row, column=2).value = finding.get('line_number', '')
            ws.cell(row=row, column=3).value = finding.get('rule_id', '')
            ws.cell(row=row, column=4).value = finding.get('message', '')
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=5).value = ""
            ws.cell(row=row, column=6).value = finding.get('suggestion', '')
            ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
            
            severity = finding.get('severity', 'low')
            ws.cell(row=row, column=7).value = severity
            ws.cell(row=row, column=7).fill = ExcelTheme.get_severity_fill(severity)
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = ExcelTheme.BORDER
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        
        ws.freeze_panes = "A2"
    
    def _create_decision_trail_sheet(self, wb: Any, report: Dict[str, Any]) -> None:
        """Create decision_trail sheet for fixer audit trail.
        
        Args:
            wb: Workbook object
            report: Compliance report
        """
        ws = wb.create_sheet("decision_trail")
        
        headers = ["Timestamp", "File", "Rule ID", "Source", "Decision", "Constraint"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = ExcelTheme.HEADER_FILL
            cell.font = ExcelTheme.HEADER_FONT
            cell.border = ExcelTheme.BORDER
        
        # Get decision trail from report
        decisions = report.get('decision_trail', [])
        
        row = 2
        for decision in decisions:
            ws.cell(row=row, column=1).value = decision.get('timestamp', '')
            ws.cell(row=row, column=2).value = decision.get('file', '')
            ws.cell(row=row, column=3).value = decision.get('rule_id', '')
            ws.cell(row=row, column=4).value = decision.get('source', '')
            ws.cell(row=row, column=5).value = decision.get('decision', '')
            ws.cell(row=row, column=6).value = decision.get('constraint', '')
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = ExcelTheme.BORDER
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        
        ws.freeze_panes = "A2"
