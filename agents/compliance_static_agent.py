"""
ORCA Compliance Static Agent — 7-phase static analysis auditor.

This module implements a comprehensive compliance auditing pipeline following
the CURE StaticAnalyzerAgent pattern, adapted for C/C++ compliance (not HDL).

Architecture:
    1. File Discovery      — Scan codebase for C/C++ source files
    2. Style Analysis      — Check code style, indentation, naming
    3. License Scan        — Validate SPDX headers and copyright
    4. Structure Check     — Validate code structure and include guards
    5. Commit Lint         — Validate commit messages and history
    6. Adapter Integration — Run external tools (checkpatch, etc.)
    7. Aggregation         — Deduplicate, score, and report findings

Features:
    - Memory-efficient batch processing with configurable batch_size
    - Try/except imports with AVAILABILITY FLAGS for graceful degradation
    - Support for both directory and single-file paths
    - Rich progress display (optional, with try/except)
    - Proper dataclasses: ComplianceReport, AnalysisMetrics
    - JSON + Excel report generation
    - HITL context support (optional)
    - Telemetry/timing tracking per phase
    - Health scoring with A-F grades
    - Domain filtering (analyze only specific domains)
"""

import os
import json
import time
import logging
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Set, Tuple
import traceback

logger = logging.getLogger(__name__)

# ───────────────────────────────────────────────────────────────────────────
# AVAILABILITY FLAGS (try/except imports with graceful degradation)
# ───────────────────────────────────────────────────────────────────────────

RICH_AVAILABLE = False
try:
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
    from rich.console import Console
    RICH_AVAILABLE = True
except ImportError:
    logger.debug("rich library not available; disabling rich progress display")

OPENPYXL_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    logger.debug("openpyxl library not available; disabling Excel export")

# ───────────────────────────────────────────────────────────────────────────
# DATACLASSES
# ───────────────────────────────────────────────────────────────────────────


@dataclass
class AnalysisMetrics:
    """Metrics for a single analysis phase."""
    phase_name: str
    duration_seconds: float
    files_processed: int
    findings_count: int
    errors_count: int
    severity_breakdown: Dict[str, int] = field(default_factory=dict)


@dataclass
class ComplianceReport:
    """Consolidated compliance audit report with detailed metadata."""
    codebase_path: str
    findings: List[Any] = field(default_factory=list)
    by_domain: Dict[str, List[Any]] = field(default_factory=dict)
    domain_scores: Dict[str, float] = field(default_factory=dict)
    domain_metrics: Dict[str, AnalysisMetrics] = field(default_factory=dict)
    overall_score: float = 1.0
    overall_grade: str = "A"
    file_count: int = 0
    lines_analyzed: int = 0
    timestamp: datetime = field(default_factory=datetime.now)
    timing: Dict[str, float] = field(default_factory=dict)
    is_compliant: bool = True
    severity_breakdown: Dict[str, int] = field(default_factory=dict)
    analysis_duration_seconds: float = 0.0

    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary for JSON serialization."""
        data = asdict(self)
        data['timestamp'] = self.timestamp.isoformat()
        return data


# ───────────────────────────────────────────────────────────────────────────
# MAIN AGENT CLASS
# ───────────────────────────────────────────────────────────────────────────


class ComplianceStaticAgent:
    """
    7-phase static analysis compliance auditor for C/C++ codebases.

    Follows CURE StaticAnalyzerAgent pattern with support for:
    - Multiple concurrent analyzers
    - External tool adapters
    - Memory-efficient batch processing
    - Configurable domain filtering
    - HITL context support
    """

    PHASES = [
        'file_discovery',
        'style_analysis',
        'license_scan',
        'structure_check',
        'commit_lint',
        'adapter_integration',
        'aggregation'
    ]

    C_EXTENSIONS = {'.c', '.h', '.cpp', '.cc', '.cxx', '.hpp', '.hxx'}

    def __init__(
        self,
        codebase_path: str,
        output_dir: str,
        config: Dict[str, Any],
        llm_tools: Optional[Any] = None,
        rules: Optional[Dict[str, Any]] = None,
        file_extensions: Optional[Set[str]] = None,
        max_files: Optional[int] = None,
        exclude_dirs: Optional[List[str]] = None,
        exclude_globs: Optional[List[str]] = None,
        batch_size: int = 50,
        memory_limit_mb: int = 512,
        enable_llm: bool = False,
        enable_adapters: bool = True,
        verbose: bool = False,
        hitl_context: Optional[Dict[str, Any]] = None,
        domains: Optional[List[str]] = None,
    ):
        """
        Initialize the compliance static agent.

        Args:
            codebase_path: Root directory or file path to analyze
            output_dir: Directory for report output
            config: Configuration dictionary
            llm_tools: Optional LLM tools instance for enrichment
            rules: Compliance rules dictionary
            file_extensions: Set of file extensions to analyze (default: C_EXTENSIONS)
            max_files: Maximum number of files to process
            exclude_dirs: Directories to exclude from analysis
            exclude_globs: Glob patterns to exclude
            batch_size: Number of files per batch for memory efficiency
            memory_limit_mb: Memory limit per batch in MB
            enable_llm: Enable LLM-based enrichment
            enable_adapters: Enable external tool adapters
            verbose: Enable verbose logging
            hitl_context: Optional HITL (Human-In-The-Loop) context
            domains: List of domains to analyze (None = all)
        """
        self.codebase_path = codebase_path
        self.output_dir = output_dir
        self.config = config or {}
        self.llm_tools = llm_tools
        self.rules = rules or {}
        self.file_extensions = file_extensions or self.C_EXTENSIONS
        self.max_files = max_files
        self.exclude_dirs = exclude_dirs or ['.git', '__pycache__', '.venv', 'node_modules']
        self.exclude_globs = exclude_globs or ['*.pyc', '.git/*', '__pycache__/*']
        self.batch_size = batch_size
        self.memory_limit_mb = memory_limit_mb
        self.enable_llm = enable_llm
        self.enable_adapters = enable_adapters
        self.verbose = verbose
        self.hitl_context = hitl_context or {}
        self.domains = domains  # None = analyze all domains

        self.analyzers: Dict[str, Any] = {}
        self.adapters: Dict[str, Any] = {}
        self.file_processor: Optional[Any] = None

        if self.verbose:
            logging.basicConfig(level=logging.DEBUG)
            logger.setLevel(logging.DEBUG)

        self._init_analyzers()
        self._init_adapters()
        self._init_file_processor()

    def _init_analyzers(self) -> None:
        """Initialize all compliance analyzers with availability flags."""
        analyzer_config = {
            'exclude_patterns': self.exclude_globs,
            'severity_overrides': self.config.get('severity_overrides', {}),
            'indent_style': self.config.get('indent_style', 'spaces'),
            'indent_size': self.config.get('indent_size', 4),
        }

        analyzers_map = {
            'style': 'agents.analyzers.style_analyzer.StyleAnalyzer',
            'license': 'agents.analyzers.license_analyzer.LicenseAnalyzer',
            'structure': 'agents.analyzers.structure_analyzer.StructureAnalyzer',
            'commit': 'agents.analyzers.commit_analyzer.CommitAnalyzer',
            'whitespace': 'agents.analyzers.whitespace_analyzer.WhitespaceAnalyzer',
            'macro': 'agents.analyzers.macro_analyzer.MacroAnalyzer',
            'include': 'agents.analyzers.include_analyzer.IncludeAnalyzer',
        }

        for analyzer_key, module_path in analyzers_map.items():
            try:
                module_name, class_name = module_path.rsplit('.', 1)
                module = __import__(module_name, fromlist=[class_name])
                analyzer_class = getattr(module, class_name)
                self.analyzers[analyzer_key] = analyzer_class(
                    self.rules.get(analyzer_key, {}),
                    analyzer_config
                )
                logger.info(f"Initialized {analyzer_key} analyzer")
            except (ImportError, AttributeError) as e:
                logger.warning(f"Could not load {analyzer_key} analyzer: {e}")

    def _init_adapters(self) -> None:
        """Initialize external tool adapters with availability flags."""
        if not self.enable_adapters:
            logger.info("Adapters disabled")
            return

        adapters_map = {
            'spdx': 'agents.adapters.spdx_adapter.SPDXAdapter',
            'include_guard': 'agents.adapters.include_guard_adapter.IncludeGuardAdapter',
            'commit_message': 'agents.adapters.commit_message_adapter.CommitMessageAdapter',
            'checkpatch': 'agents.adapters.checkpatch_adapter.CheckpatchAdapter',
        }

        for adapter_key, module_path in adapters_map.items():
            try:
                module_name, class_name = module_path.rsplit('.', 1)
                module = __import__(module_name, fromlist=[class_name])
                adapter_class = getattr(module, class_name)
                self.adapters[adapter_key] = adapter_class(
                    self.rules.get(adapter_key, {}),
                    self.config
                )
                logger.info(f"Initialized {adapter_key} adapter")
            except (ImportError, AttributeError) as e:
                logger.warning(f"Could not load {adapter_key} adapter: {e}")

    def _init_file_processor(self) -> None:
        """Initialize the file processor."""
        try:
            from agents.core.file_processor import FileProcessor
            processor_config = {
                'exclude_patterns': self.exclude_globs,
                'file_extensions': self.file_extensions,
            }
            self.file_processor = FileProcessor(processor_config)
            logger.info("Initialized FileProcessor")
        except ImportError as e:
            logger.warning(f"Could not import FileProcessor: {e}")

    def run_audit(self) -> ComplianceReport:
        """
        Execute the 7-phase compliance audit pipeline.

        Returns:
            ComplianceReport with all findings and scoring
        """
        logger.info(f"Starting compliance audit on {self.codebase_path}")
        audit_start = time.time()

        report = ComplianceReport(codebase_path=self.codebase_path)

        try:
            # Phase 1: File Discovery
            discovered_files = self._phase_1_discover_files()
            report.file_count = len(discovered_files)
            logger.info(f"Phase 1: Discovered {report.file_count} files")

            if not discovered_files:
                logger.warning("No files to analyze")
                return report

            # Phase 2: Style Analysis
            if self._should_analyze_domain('style'):
                self._phase_2_style_analysis(discovered_files, report)

            # Phase 3: License Scan
            if self._should_analyze_domain('license'):
                self._phase_3_license_scan(discovered_files, report)

            # Phase 4: Structure Check
            if self._should_analyze_domain('structure'):
                self._phase_4_structure_check(discovered_files, report)

            # Phase 5: Commit Lint
            if self._should_analyze_domain('commits'):
                self._phase_5_commit_lint(discovered_files, report)

            # Phase 6: Adapter Integration
            if self.enable_adapters:
                self._phase_6_adapter_integration(discovered_files, report)

            # Phase 7: Aggregation
            self._phase_7_aggregation(report)

            # Calculate scores and grades
            self._calculate_scores(report)

        except Exception as e:
            logger.error(f"Fatal error during audit: {e}")
            logger.debug(traceback.format_exc())
            report.is_compliant = False
            report.overall_grade = "F"

        finally:
            report.analysis_duration_seconds = time.time() - audit_start
            logger.info(
                f"Audit complete in {report.analysis_duration_seconds:.2f}s. "
                f"Findings: {len(report.findings)}, Grade: {report.overall_grade}"
            )

        return report

    # ───────────────────────────────────────────────────────────────────────
    # PHASE IMPLEMENTATIONS
    # ───────────────────────────────────────────────────────────────────────

    def _phase_1_discover_files(self) -> List[str]:
        """
        Phase 1: File Discovery.

        Accepts both directory (recursive walk) and single-file paths.
        """
        phase_start = time.time()
        logger.info(f"Phase 1: File Discovery on {self.codebase_path}")

        files = []

        # Single-file mode
        if os.path.isfile(self.codebase_path):
            logger.info(f"Single-file mode: {self.codebase_path}")
            files.append(self.codebase_path)
            return files

        # Directory mode with optional FileProcessor
        if self.file_processor:
            try:
                from agents.core.file_processor import FileMetadata
                metadata_list = self.file_processor.discover_files(self.codebase_path)
                files = [m.path for m in metadata_list]
            except Exception as e:
                logger.warning(f"FileProcessor failed, falling back to os.walk: {e}")
                files = self._discover_files_fallback()
        else:
            files = self._discover_files_fallback()

        # Apply max_files limit
        if self.max_files and len(files) > self.max_files:
            logger.warning(f"Limiting to {self.max_files} files (found {len(files)})")
            files = files[:self.max_files]

        elapsed = time.time() - phase_start
        logger.info(f"Phase 1 complete in {elapsed:.2f}s: discovered {len(files)} files")

        return files

    def _discover_files_fallback(self) -> List[str]:
        """Fallback file discovery using os.walk."""
        files = []
        for root, dirs, filenames in os.walk(self.codebase_path):
            dirs[:] = [d for d in dirs if d not in self.exclude_dirs]

            for filename in filenames:
                _, ext = os.path.splitext(filename)
                if ext in self.file_extensions:
                    files.append(os.path.join(root, filename))

        return files

    def _phase_2_style_analysis(
        self, discovered_files: List[str], report: ComplianceReport
    ) -> None:
        """Phase 2: Style Analysis (indentation, naming, spacing)."""
        phase_start = time.time()
        logger.info("Phase 2: Style Analysis")

        phase_findings = self._run_analyzers_batch(
            discovered_files,
            ['style', 'whitespace', 'macro']
        )

        self._add_findings_to_report(report, phase_findings, 'style')
        elapsed = time.time() - phase_start
        report.timing['phase_2_style'] = elapsed
        logger.info(f"Phase 2 complete in {elapsed:.2f}s: {len(phase_findings)} findings")

    def _phase_3_license_scan(
        self, discovered_files: List[str], report: ComplianceReport
    ) -> None:
        """Phase 3: License Scan (SPDX headers, copyright)."""
        phase_start = time.time()
        logger.info("Phase 3: License Scan")

        phase_findings = self._run_analyzers_batch(
            discovered_files,
            ['license']
        )

        self._add_findings_to_report(report, phase_findings, 'license')

        # Run SPDX adapter if available
        if 'spdx' in self.adapters:
            try:
                adapter_findings = self._run_adapter('spdx', discovered_files)
                self._add_findings_to_report(report, adapter_findings, 'license')
            except Exception as e:
                logger.warning(f"SPDX adapter failed: {e}")

        elapsed = time.time() - phase_start
        report.timing['phase_3_license'] = elapsed
        logger.info(f"Phase 3 complete in {elapsed:.2f}s: {len(phase_findings)} findings")

    def _phase_4_structure_check(
        self, discovered_files: List[str], report: ComplianceReport
    ) -> None:
        """Phase 4: Structure Check (include guards, file layout)."""
        phase_start = time.time()
        logger.info("Phase 4: Structure Check")

        phase_findings = self._run_analyzers_batch(
            discovered_files,
            ['structure', 'include']
        )

        self._add_findings_to_report(report, phase_findings, 'structure')

        # Run include guard adapter if available
        if 'include_guard' in self.adapters:
            try:
                adapter_findings = self._run_adapter('include_guard', discovered_files)
                self._add_findings_to_report(report, adapter_findings, 'structure')
            except Exception as e:
                logger.warning(f"Include guard adapter failed: {e}")

        elapsed = time.time() - phase_start
        report.timing['phase_4_structure'] = elapsed
        logger.info(f"Phase 4 complete in {elapsed:.2f}s: {len(phase_findings)} findings")

    def _phase_5_commit_lint(
        self, discovered_files: List[str], report: ComplianceReport
    ) -> None:
        """Phase 5: Commit Lint (commit message format, history)."""
        phase_start = time.time()
        logger.info("Phase 5: Commit Lint")

        phase_findings = self._run_analyzers_batch(
            discovered_files,
            ['commit']
        )

        self._add_findings_to_report(report, phase_findings, 'commits')

        # Run commit message adapter if available
        if 'commit_message' in self.adapters:
            try:
                adapter_findings = self._run_adapter('commit_message', discovered_files)
                self._add_findings_to_report(report, adapter_findings, 'commits')
            except Exception as e:
                logger.warning(f"Commit message adapter failed: {e}")

        elapsed = time.time() - phase_start
        report.timing['phase_5_commit'] = elapsed
        logger.info(f"Phase 5 complete in {elapsed:.2f}s: {len(phase_findings)} findings")

    def _phase_6_adapter_integration(
        self, discovered_files: List[str], report: ComplianceReport
    ) -> None:
        """Phase 6: Adapter Integration (checkpatch, external tools)."""
        phase_start = time.time()
        logger.info("Phase 6: Adapter Integration")

        # Run checkpatch adapter if available
        if 'checkpatch' in self.adapters:
            try:
                adapter_findings = self._run_adapter('checkpatch', discovered_files)
                self._add_findings_to_report(report, adapter_findings, 'checkpatch')
            except Exception as e:
                logger.warning(f"Checkpatch adapter failed: {e}")

        elapsed = time.time() - phase_start
        report.timing['phase_6_adapters'] = elapsed
        logger.info(f"Phase 6 complete in {elapsed:.2f}s")

    def _phase_7_aggregation(self, report: ComplianceReport) -> None:
        """Phase 7: Aggregation (deduplication, grouping, summary)."""
        phase_start = time.time()
        logger.info("Phase 7: Aggregation")

        # Deduplicate findings
        unique_findings = {}
        for finding in report.findings:
            key = self._finding_key(finding)
            if key not in unique_findings:
                unique_findings[key] = finding

        report.findings = list(unique_findings.values())

        # Group by domain
        for finding in report.findings:
            domain = self._get_finding_field(finding, 'category', 'unknown')
            if domain not in report.by_domain:
                report.by_domain[domain] = []
            report.by_domain[domain].append(finding)

        # Calculate severity breakdown
        report.severity_breakdown = {}
        for finding in report.findings:
            severity = self._get_finding_field(finding, 'severity', 'UNKNOWN')
            report.severity_breakdown[severity] = report.severity_breakdown.get(severity, 0) + 1

        elapsed = time.time() - phase_start
        report.timing['phase_7_aggregation'] = elapsed
        logger.info(
            f"Phase 7 complete in {elapsed:.2f}s: "
            f"{len(report.findings)} unique findings, "
            f"{len(report.by_domain)} domains"
        )

    # ───────────────────────────────────────────────────────────────────────
    # HELPER METHODS
    # ───────────────────────────────────────────────────────────────────────

    def _should_analyze_domain(self, domain: str) -> bool:
        """Check if a domain should be analyzed based on domain filter."""
        if self.domains is None:
            return True
        return domain in self.domains

    def _run_analyzers_batch(
        self, files: List[str], analyzer_keys: List[str]
    ) -> List[Any]:
        """
        Run analyzers in memory-efficient batches.

        Args:
            files: List of file paths to analyze
            analyzer_keys: List of analyzer keys to run

        Returns:
            List of findings from all analyzers
        """
        all_findings = []

        # Process files in batches
        for batch_start in range(0, len(files), self.batch_size):
            batch_end = min(batch_start + self.batch_size, len(files))
            batch = files[batch_start:batch_end]

            logger.debug(f"Processing batch {batch_start // self.batch_size + 1}: {len(batch)} files")

            for analyzer_key in analyzer_keys:
                if analyzer_key not in self.analyzers:
                    continue

                analyzer = self.analyzers[analyzer_key]
                for file_path in batch:
                    try:
                        if not os.path.isfile(file_path):
                            continue

                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()

                        findings = analyzer.analyze(file_path, content)
                        if findings:
                            all_findings.extend(findings)

                    except Exception as e:
                        logger.warning(f"Error analyzing {file_path} with {analyzer_key}: {e}")

        return all_findings

    def _run_adapter(self, adapter_key: str, files: List[str]) -> List[Any]:
        """
        Run an adapter on files.

        Args:
            adapter_key: Key of the adapter to run
            files: List of file paths

        Returns:
            List of findings from the adapter
        """
        if adapter_key not in self.adapters:
            return []

        adapter = self.adapters[adapter_key]
        findings = []

        try:
            # Build file cache for batch analysis
            file_cache = {}
            for file_path in files[:self.batch_size]:  # Limit to batch size
                if os.path.isfile(file_path):
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            file_cache[file_path] = f.read()
                    except Exception as e:
                        logger.debug(f"Could not read {file_path}: {e}")

            if not file_cache:
                return []

            # Run adapter
            result = adapter.analyze(file_cache)

            # Convert result to findings list
            if hasattr(result, 'findings'):
                findings = result.findings
            elif isinstance(result, list):
                findings = result
            else:
                logger.warning(f"Unexpected adapter result type: {type(result)}")

        except Exception as e:
            logger.warning(f"Adapter {adapter_key} failed: {e}")

        return findings

    def _add_findings_to_report(
        self, report: ComplianceReport, findings: List[Any], domain: str
    ) -> None:
        """Add findings to report and update domain tracking."""
        report.findings.extend(findings)
        if domain not in report.by_domain:
            report.by_domain[domain] = []
        report.by_domain[domain].extend(findings)

    def _finding_key(self, finding: Any) -> Tuple[str, int, str]:
        """Generate deduplication key for a finding."""
        file_path = self._get_finding_field(finding, 'file_path', '')
        line_number = self._get_finding_field(finding, 'line_number', 0)
        message = self._get_finding_field(finding, 'message', '')
        return (file_path, line_number, message)

    def _get_finding_field(self, finding: Any, field_name: str, default: Any = None) -> Any:
        """Get field from finding (handles both dict and dataclass)."""
        if isinstance(finding, dict):
            return finding.get(field_name, default)
        else:
            return getattr(finding, field_name, default)

    def _calculate_scores(self, report: ComplianceReport) -> None:
        """
        Calculate compliance scores and grades.

        Updates report in-place with:
        - overall_score (0-1)
        - overall_grade (A-F)
        - domain_scores
        """
        if not report.findings:
            report.overall_score = 1.0
            report.overall_grade = "A"
            report.domain_scores = {}
            return

        # Severity weights
        severity_weights = {
            'CRITICAL': 0.4,
            'HIGH': 0.3,
            'MEDIUM': 0.2,
            'LOW': 0.1,
        }

        # Calculate overall score from findings
        total_weight = sum(
            severity_weights.get(
                self._get_finding_field(f, 'severity', 'LOW'), 0.0
            )
            for f in report.findings
        )

        # Normalize to 0-1 (lower weight = higher score)
        report.overall_score = max(0.0, 1.0 - (total_weight * 0.05))

        # Calculate domain scores
        for domain, findings in report.by_domain.items():
            domain_weight = sum(
                severity_weights.get(
                    self._get_finding_field(f, 'severity', 'LOW'), 0.0
                )
                for f in findings
            )
            domain_score = max(0.0, 1.0 - (domain_weight * 0.05))
            report.domain_scores[domain] = domain_score

        # Compute grade from score
        report.overall_grade = self._compute_grade(report.overall_score)
        report.is_compliant = (
            report.overall_score >= self.config.get('compliance_threshold', 0.8)
        )

    def _compute_grade(self, score: float) -> str:
        """
        Compute letter grade from numerical score.

        Args:
            score: Score from 0.0 to 1.0

        Returns:
            Grade letter (A-F)
        """
        if score >= 0.9:
            return "A"
        elif score >= 0.8:
            return "B"
        elif score >= 0.7:
            return "C"
        elif score >= 0.6:
            return "D"
        else:
            return "F"

    # ───────────────────────────────────────────────────────────────────────
    # REPORT GENERATION
    # ───────────────────────────────────────────────────────────────────────

    def generate_reports(self, report: ComplianceReport) -> Dict[str, str]:
        """
        Generate all report formats (JSON, Excel).

        Args:
            report: Compliance report to export

        Returns:
            Dictionary mapping format -> file path
        """
        os.makedirs(self.output_dir, exist_ok=True)
        generated = {}

        # JSON report
        try:
            json_path = self._write_json_report(report)
            generated['json'] = json_path
            logger.info(f"JSON report written to {json_path}")
        except Exception as e:
            logger.error(f"Failed to write JSON report: {e}")

        # Excel report
        if OPENPYXL_AVAILABLE:
            try:
                excel_path = self._write_excel_report(report)
                generated['excel'] = excel_path
                logger.info(f"Excel report written to {excel_path}")
            except Exception as e:
                logger.error(f"Failed to write Excel report: {e}")
        else:
            logger.debug("openpyxl not available; skipping Excel report")

        return generated

    def _write_json_report(self, report: ComplianceReport) -> str:
        """
        Write JSON report to file.

        Args:
            report: Compliance report

        Returns:
            Path to written file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        json_path = os.path.join(self.output_dir, f'compliance_report_{timestamp}.json')

        data = report.to_dict()

        # Convert findings to serializable format
        findings_list = []
        for finding in data.get('findings', []):
            if isinstance(finding, dict):
                findings_list.append(finding)
            else:
                # Convert dataclass to dict
                finding_dict = asdict(finding) if hasattr(finding, '__dataclass_fields__') else {
                    'file_path': str(getattr(finding, 'file_path', '')),
                    'line_number': getattr(finding, 'line_number', 0),
                    'severity': getattr(finding, 'severity', 'UNKNOWN'),
                    'category': getattr(finding, 'category', 'unknown'),
                    'message': getattr(finding, 'message', ''),
                }
                findings_list.append(finding_dict)

        data['findings'] = findings_list

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

        return json_path

    def _write_excel_report(self, report: ComplianceReport) -> str:
        """
        Write Excel report to file.

        Args:
            report: Compliance report

        Returns:
            Path to written file
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel export")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = os.path.join(self.output_dir, f'compliance_report_{timestamp}.xlsx')

        wb = Workbook()
        wb.remove(wb.active)

        # Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary['A1'] = "Compliance Audit Report"
        ws_summary['A2'] = f"Codebase: {report.codebase_path}"
        ws_summary['A3'] = f"Timestamp: {report.timestamp.isoformat()}"
        ws_summary['A4'] = f"Overall Grade: {report.overall_grade}"
        ws_summary['A5'] = f"Overall Score: {report.overall_score:.1%}"
        ws_summary['A6'] = f"Files Analyzed: {report.file_count}"
        ws_summary['A7'] = f"Total Findings: {len(report.findings)}"

        # Findings sheet
        ws_findings = wb.create_sheet("Findings", 1)
        headers = ['File', 'Line', 'Severity', 'Category', 'Message']
        for col, header in enumerate(headers, 1):
            ws_findings.cell(row=1, column=col, value=header)

        row = 2
        for finding in report.findings:
            ws_findings.cell(row=row, column=1, value=self._get_finding_field(finding, 'file_path', ''))
            ws_findings.cell(row=row, column=2, value=self._get_finding_field(finding, 'line_number', ''))
            ws_findings.cell(row=row, column=3, value=self._get_finding_field(finding, 'severity', ''))
            ws_findings.cell(row=row, column=4, value=self._get_finding_field(finding, 'category', ''))
            ws_findings.cell(row=row, column=5, value=self._get_finding_field(finding, 'message', ''))
            row += 1

        wb.save(excel_path)
        return excel_path
