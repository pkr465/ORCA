"""
Comprehensive fix recommendation generator for compliance findings using LLM.

This module generates detailed, confidence-scored fix recommendations for compliance
findings, incorporating human-in-the-loop context, constraint rules, and alternative
solutions. Uses utils.llm_tools.LLMTools for all LLM interactions.
"""

import json
import logging
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.llm_tools import LLMTools
from agents.prompts.solution_prompts import SOLUTION_SYSTEM_PROMPT, SOLUTION_USER_PROMPT_TEMPLATE

logger = logging.getLogger(__name__)


@dataclass
class Solution:
    """Represents a fix solution for a compliance finding."""

    finding_id: str
    fix_type: str
    domain: str
    file_path: str
    line_number: int
    old_code: str
    new_code: str
    explanation: str
    confidence: float
    confidence_level: str = field(default="LOW")
    alternatives: List[Dict[str, Any]] = field(default_factory=list)
    side_effects: List[Dict[str, str]] = field(default_factory=list)
    requires_manual_review: bool = False
    review_notes: str = ""
    root_cause: str = ""
    constraints_applied: List[str] = field(default_factory=list)

    def __post_init__(self):
        """Set confidence_level based on confidence score."""
        if self.confidence >= 0.8:
            self.confidence_level = "HIGH"
        elif self.confidence >= 0.5:
            self.confidence_level = "MEDIUM"
        else:
            self.confidence_level = "LOW"

    def to_dict(self) -> Dict[str, Any]:
        """Convert solution to dictionary for serialization."""
        return {
            "finding_id": self.finding_id,
            "fix_type": self.fix_type,
            "domain": self.domain,
            "file_path": self.file_path,
            "line_number": self.line_number,
            "old_code": self.old_code,
            "new_code": self.new_code,
            "explanation": self.explanation,
            "confidence": self.confidence,
            "confidence_level": self.confidence_level,
            "alternatives": self.alternatives,
            "side_effects": self.side_effects,
            "requires_manual_review": self.requires_manual_review,
            "review_notes": self.review_notes,
            "root_cause": self.root_cause,
            "constraints_applied": self.constraints_applied,
        }


class ComplianceSolutionAgent:
    """
    Generates fix recommendations for compliance findings using LLM.

    Features:
    - LLM-powered fix generation with confidence scoring
    - Code context extraction (±10 lines around findings)
    - Human-in-the-loop context injection from past decisions
    - Constraint-based decision guidance from markdown files
    - Alternative solution generation
    - Side effect assessment
    - Batch processing by file
    - Solution validation
    - Excel report writing
    - Telemetry tracking
    """

    # Valid fix types
    VALID_FIX_TYPES = {
        "SIMPLE_REPLACEMENT",
        "ADD_CODE",
        "REMOVE_CODE",
        "REFACTOR",
        "STRUCTURAL",
        "REQUIRES_MANUAL_REVIEW",
    }

    def __init__(
        self,
        config: Dict[str, Any],
        llm_tools: LLMTools,
        rules: Dict[str, Any],
        hitl_context: Optional[Dict[str, Any]] = None,
        constraints_dir: Optional[str] = None,
        telemetry: Optional[Any] = None,
        telemetry_run_id: Optional[str] = None,
    ):
        """
        Initialize the ComplianceSolutionAgent.

        Args:
            config: Configuration dictionary
            llm_tools: LLMTools instance for LLM interactions
            rules: Compliance rules dictionary
            hitl_context: Human-in-the-loop context with past decisions
            constraints_dir: Directory containing constraint markdown files
            telemetry: Telemetry tracker instance
            telemetry_run_id: Run ID for telemetry tracking
        """
        self.config = config
        self.llm_tools = llm_tools
        self.rules = rules
        self.hitl_context = hitl_context or {}
        self.constraints_dir = constraints_dir
        self.telemetry = telemetry
        self.telemetry_run_id = telemetry_run_id
        self.context_lines = config.get("context_lines", 10)
        self.max_retries = config.get("max_retries", 2)
        self.temperature = config.get("temperature", 0.3)

        # Cache for loaded constraints
        self._constraints_cache: Dict[str, str] = {}

        logger.info("ComplianceSolutionAgent initialized with LLMTools")

    def generate_solutions(
        self, findings: List[Dict[str, Any]], file_contents: Dict[str, str]
    ) -> Dict[str, Solution]:
        """
        Generate solutions for all findings, grouped and processed by file.

        Args:
            findings: List of finding dictionaries from compliance analysis
            file_contents: Dictionary mapping file paths to file contents

        Returns:
            Dictionary mapping finding IDs to Solution objects
        """
        logger.info(f"Generating solutions for {len(findings)} findings")
        solutions = {}

        # Group findings by file for batch processing
        findings_by_file: Dict[str, List[Dict[str, Any]]] = {}
        for finding in findings:
            file_path = finding.get("file_path", "")
            if file_path not in findings_by_file:
                findings_by_file[file_path] = []
            findings_by_file[file_path].append(finding)

        # Process each file
        for file_path, file_findings in findings_by_file.items():
            logger.debug(f"Processing {len(file_findings)} findings in {file_path}")

            if file_path not in file_contents:
                logger.warning(f"File content not available for {file_path}")
                continue

            content = file_contents[file_path]

            # Generate solutions for each finding in the file
            for finding in file_findings:
                try:
                    finding_id = finding.get("id", "")
                    solution = self._generate_solution(finding, content)
                    solutions[finding_id] = solution

                    # Track telemetry
                    if self.telemetry and self.telemetry_run_id:
                        self.telemetry.track_event(
                            "solution_generated",
                            {
                                "run_id": self.telemetry_run_id,
                                "finding_id": finding_id,
                                "confidence": solution.confidence,
                                "fix_type": solution.fix_type,
                            },
                        )

                except Exception as e:
                    logger.error(
                        f"Error generating solution for finding {finding.get('id')}: {e}",
                        exc_info=True,
                    )
                    # Create fallback solution
                    solutions[finding.get("id", "")] = self._create_fallback_solution(
                        finding
                    )

        logger.info(f"Generated {len(solutions)} solutions")
        return solutions

    def _generate_solution(self, finding: Dict[str, Any], file_content: str) -> Solution:
        """
        Generate a single solution using LLM.

        Args:
            finding: Finding dictionary
            file_content: File content as string

        Returns:
            Solution object with fix recommendations

        Raises:
            Exception: On LLM errors (fallback solution created upstream)
        """
        logger.debug(f"Generating solution for finding: {finding.get('message')}")

        finding_id = finding.get("id", "")
        domain = finding.get("domain", "style")

        # Build code context with line markers
        context = self._build_context(finding, file_content, self.context_lines)

        # Load constraint rules for this domain
        constraints = self._load_constraints(domain)

        # Build complete prompt
        prompt = self._build_solution_prompt(finding, context, domain, constraints)

        # Inject HITL context from past decisions
        prompt = self._inject_hitl_context(prompt, finding)

        # Call LLM for solution generation
        try:
            response = self.llm_tools.chat_completion(
                messages=[
                    {"role": "system", "content": SOLUTION_SYSTEM_PROMPT},
                    {"role": "user", "content": prompt},
                ],
                temperature=self.temperature,
                max_tokens=4096,
            )

            # Parse response
            solution = self._parse_solution_response(response.content, finding)

            # Validate solution
            if self._validate_solution(solution, file_content):
                logger.debug(f"Solution validated for {finding_id}")
            else:
                logger.warning(f"Solution validation failed for {finding_id}")
                solution.requires_manual_review = True
                solution.review_notes += " (Validation failed)"

            return solution

        except Exception as e:
            logger.error(f"LLM solution generation failed: {e}", exc_info=True)
            return self._create_fallback_solution(finding)

    def _build_context(
        self, finding: Dict[str, Any], file_content: str, context_lines: int = 10
    ) -> str:
        """
        Build surrounding code context with line markers.

        Args:
            finding: Finding dictionary with line_number
            file_content: Full file content as string
            context_lines: Number of lines to show before/after

        Returns:
            Formatted code context with line markers
        """
        lines = file_content.split("\n")
        line_num = finding.get("line_number", 1) - 1

        start = max(0, line_num - context_lines)
        end = min(len(lines), line_num + context_lines + 1)

        context_lines_list = []
        for i in range(start, end):
            marker = ">>>" if i == line_num else "   "
            context_lines_list.append(f"{marker} {i + 1:4d}: {lines[i]}")

        return "\n".join(context_lines_list)

    def _inject_hitl_context(self, prompt: str, finding: Dict[str, Any]) -> str:
        """
        Inject human-in-the-loop decision context into prompt.

        Args:
            prompt: Base prompt string
            finding: Current finding being analyzed

        Returns:
            Prompt with HITL context injected
        """
        similar_decisions = []
        finding_type = finding.get("finding_type", "")
        domain = finding.get("domain", "")

        # Find similar past decisions
        for decision in self.hitl_context.get("decisions", []):
            if (
                decision.get("finding_type") == finding_type
                or decision.get("domain") == domain
            ):
                similar_decisions.append(decision)

        if similar_decisions:
            hitl_section = "\n\nHUMAN-IN-THE-LOOP CONTEXT (Previous similar decisions):\n"
            for i, decision in enumerate(similar_decisions[:3], 1):
                hitl_section += f"{i}. Finding Type: {decision.get('finding_type')}\n"
                hitl_section += f"   Decision: {decision.get('decision')}\n"
                hitl_section += f"   Rationale: {decision.get('rationale', 'N/A')}\n"
            prompt += hitl_section

        return prompt

    def _load_constraints(self, domain: str) -> str:
        """
        Load constraint rules from markdown files for a domain.

        Loads common/sample constraints first (applies to all domains),
        then domain-specific constraints, and concatenates both.

        Args:
            domain: Domain name (e.g., 'style', 'license', 'structure')

        Returns:
            Formatted constraint rules as string
        """
        if domain in self._constraints_cache:
            return self._constraints_cache[domain]

        if not self.constraints_dir or not Path(self.constraints_dir).is_dir():
            return ""

        parts = []
        base = Path(self.constraints_dir)

        # 1. Load common/sample constraints (applies to ALL domains)
        for name in ("common_constraints.md", "sample_constraints.md"):
            common_file = base / name
            if common_file.exists():
                try:
                    parts.append(common_file.read_text(encoding="utf-8"))
                except Exception as e:
                    logger.warning(f"Could not load {name}: {e}")
                break  # use first found

        # 2. Load domain-specific constraints
        for name in (f"{domain}_constraints.md", f"{domain}.md"):
            domain_file = base / name
            if domain_file.exists():
                try:
                    parts.append(domain_file.read_text(encoding="utf-8"))
                except Exception as e:
                    logger.warning(f"Could not load {name}: {e}")
                break

        content = "\n\n".join(parts)
        if content:
            self._constraints_cache[domain] = content
            logger.info(f"Loaded constraints for domain '{domain}' ({len(parts)} file(s))")
        else:
            logger.debug(f"No constraint files found for domain '{domain}'")

        return content

    def _build_solution_prompt(
        self,
        finding: Dict[str, Any],
        context: str,
        domain: str,
        constraints: str,
    ) -> str:
        """
        Build complete solution prompt for LLM.

        Args:
            finding: Finding dictionary
            context: Code context string
            domain: Domain name
            constraints: Constraint rules string

        Returns:
            Formatted user prompt for LLM
        """
        rules_for_domain = self.rules.get(domain, {})
        rules_str = json.dumps(rules_for_domain, indent=2)

        prompt = f"""Generate a fix recommendation for this compliance finding:

FINDING:
ID: {finding.get('id', 'N/A')}
Type: {finding.get('finding_type', 'N/A')}
Message: {finding.get('message', '')}
Severity: {finding.get('severity', 'medium')}
File: {finding.get('file_path', 'N/A')}
Line: {finding.get('line_number', 0)}

CODE CONTEXT (surrounding code with >>> marking the problematic line):
```
{context}
```

APPLICABLE RULES FOR {domain.upper()}:
{rules_str}
"""

        if constraints:
            prompt += f"\nCONSTRAINT RULES:\n{constraints}\n"

        prompt += """
Please provide a comprehensive fix recommendation:

1. Root cause analysis
2. Recommended fix type (SIMPLE_REPLACEMENT, ADD_CODE, REMOVE_CODE, REFACTOR, STRUCTURAL, or REQUIRES_MANUAL_REVIEW)
3. Exact old code snippet (what to replace) or null if adding new code
4. Exact new code snippet (replacement or new code)
5. Line range affected
6. Confidence in this solution (0.0-1.0)
7. Alternative solutions if applicable
8. Potential side effects or risks
9. Whether manual review is required

Return as JSON."""

        return prompt

    def _parse_solution_response(self, response: str, finding: Dict[str, Any]) -> Solution:
        """
        Parse LLM solution response into Solution object.

        Args:
            response: LLM response string
            finding: Original finding dictionary

        Returns:
            Solution object parsed from response
        """
        try:
            # Extract JSON from response
            json_start = response.find("{")
            json_end = response.rfind("}") + 1

            if json_start < 0 or json_end <= json_start:
                logger.warning(f"No JSON found in response for {finding.get('id')}")
                return self._create_fallback_solution(finding)

            json_str = response[json_start:json_end]
            data = json.loads(json_str)

            # Extract fields with validation
            fix_type = data.get("fix_type", "REQUIRES_MANUAL_REVIEW")
            if fix_type not in self.VALID_FIX_TYPES:
                logger.warning(f"Invalid fix_type: {fix_type}, defaulting to REQUIRES_MANUAL_REVIEW")
                fix_type = "REQUIRES_MANUAL_REVIEW"

            confidence = float(data.get("confidence", 0.5))
            confidence = max(0.0, min(1.0, confidence))  # Clamp to [0, 1]

            # Parse alternatives
            alternatives = []
            for alt in data.get("alternative_solutions", []):
                alternatives.append(
                    {
                        "description": alt.get("description", ""),
                        "old_code": alt.get("old_code", ""),
                        "new_code": alt.get("new_code", ""),
                        "confidence": float(alt.get("confidence", 0.5)),
                    }
                )

            # Parse side effects
            side_effects = []
            for effect in data.get("side_effects", []):
                side_effects.append(
                    {
                        "area": effect.get("area", ""),
                        "description": effect.get("description", ""),
                        "severity": effect.get("severity", "LOW"),
                    }
                )

            # Determine if manual review is needed
            requires_manual = data.get("requires_manual_review", False)
            if confidence < 0.5 or fix_type == "REQUIRES_MANUAL_REVIEW":
                requires_manual = True

            solution = Solution(
                finding_id=finding.get("id", ""),
                fix_type=fix_type,
                domain=finding.get("domain", ""),
                file_path=finding.get("file_path", ""),
                line_number=data.get("line_start", finding.get("line_number", 0)),
                old_code=data.get("old_code", ""),
                new_code=data.get("new_code", ""),
                explanation=data.get("explanation", ""),
                confidence=confidence,
                alternatives=alternatives,
                side_effects=side_effects,
                requires_manual_review=requires_manual,
                review_notes=data.get("review_notes", ""),
                root_cause=data.get("root_cause", ""),
                constraints_applied=data.get("constraints_applied", []),
            )

            return solution

        except json.JSONDecodeError as e:
            logger.warning(f"JSON decode error for {finding.get('id')}: {e}")
            return self._create_fallback_solution(finding)
        except Exception as e:
            logger.error(f"Error parsing solution response: {e}", exc_info=True)
            return self._create_fallback_solution(finding)

    def _validate_solution(self, solution: Solution, file_content: str) -> bool:
        """
        Validate solution against file content.

        Args:
            solution: Solution object to validate
            file_content: File content string

        Returns:
            True if solution is valid, False otherwise
        """
        # Check if old_code matches file content (for replacement solutions)
        if solution.old_code and solution.fix_type == "SIMPLE_REPLACEMENT":
            if solution.old_code not in file_content:
                logger.debug(
                    f"Old code snippet not found in file for {solution.finding_id}"
                )
                return False

        # Check line number is valid
        lines = file_content.split("\n")
        if solution.line_number < 1 or solution.line_number > len(lines):
            logger.debug(
                f"Invalid line number {solution.line_number} for {solution.finding_id}"
            )
            return False

        # Check confidence is in valid range
        if not (0.0 <= solution.confidence <= 1.0):
            logger.debug(
                f"Invalid confidence {solution.confidence} for {solution.finding_id}"
            )
            return False

        return True

    def _create_fallback_solution(self, finding: Dict[str, Any]) -> Solution:
        """
        Create a fallback solution when LLM generation fails.

        Args:
            finding: Finding dictionary

        Returns:
            Fallback Solution object
        """
        return Solution(
            finding_id=finding.get("id", ""),
            fix_type="REQUIRES_MANUAL_REVIEW",
            domain=finding.get("domain", ""),
            file_path=finding.get("file_path", ""),
            line_number=finding.get("line_number", 0),
            old_code="",
            new_code="",
            explanation="Automated solution generation failed. Manual review required.",
            confidence=0.0,
            requires_manual_review=True,
            review_notes="LLM solution generation failed",
        )

    def _write_solutions_report(self, solutions: Dict[str, Solution], output_path: str) -> None:
        """
        Write solutions to an Excel report.

        Args:
            solutions: Dictionary mapping finding IDs to Solution objects
            output_path: Path to write Excel file
        """
        logger.info(f"Writing solutions report to {output_path}")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Solutions"

            # Define headers
            headers = [
                "Finding ID",
                "File Path",
                "Line Number",
                "Domain",
                "Fix Type",
                "Confidence",
                "Confidence Level",
                "Root Cause",
                "Old Code",
                "New Code",
                "Explanation",
                "Alternative Solutions",
                "Side Effects",
                "Requires Manual Review",
                "Review Notes",
                "Constraints Applied",
            ]

            # Write headers with styling
            header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Write data rows
            for row, (finding_id, solution) in enumerate(solutions.items(), 2):
                ws.cell(row=row, column=1).value = finding_id
                ws.cell(row=row, column=2).value = solution.file_path
                ws.cell(row=row, column=3).value = solution.line_number
                ws.cell(row=row, column=4).value = solution.domain
                ws.cell(row=row, column=5).value = solution.fix_type
                ws.cell(row=row, column=6).value = f"{solution.confidence:.2f}"
                ws.cell(row=row, column=7).value = solution.confidence_level
                ws.cell(row=row, column=8).value = solution.root_cause
                ws.cell(row=row, column=9).value = solution.old_code
                ws.cell(row=row, column=10).value = solution.new_code
                ws.cell(row=row, column=11).value = solution.explanation
                ws.cell(row=row, column=12).value = json.dumps(solution.alternatives)
                ws.cell(row=row, column=13).value = json.dumps(solution.side_effects)
                ws.cell(row=row, column=14).value = "Yes" if solution.requires_manual_review else "No"
                ws.cell(row=row, column=15).value = solution.review_notes
                ws.cell(row=row, column=16).value = ", ".join(solution.constraints_applied)

            # Adjust column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 18
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 15
            ws.column_dimensions["H"].width = 25
            ws.column_dimensions["I"].width = 25
            ws.column_dimensions["J"].width = 25
            ws.column_dimensions["K"].width = 30
            ws.column_dimensions["L"].width = 20
            ws.column_dimensions["M"].width = 20
            ws.column_dimensions["N"].width = 18
            ws.column_dimensions["O"].width = 25
            ws.column_dimensions["P"].width = 20

            # Save workbook
            wb.save(output_path)
            logger.info(f"Solutions report written successfully to {output_path}")

        except Exception as e:
            logger.error(f"Error writing solutions report: {e}", exc_info=True)
            raise
