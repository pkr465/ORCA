"""Excel workbook creation and formatting utilities."""

from dataclasses import dataclass
from typing import List, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class ExcelTheme:
    """Theme configuration for Excel workbooks."""
    header_fill: str = "1B3A5C"
    header_font_color: str = "FFFFFF"
    alt_row_fill: str = "EBF2FA"
    accent_color: str = "2E75B6"
    warning_color: str = "FFC7CE"
    error_color: str = "FF6B6B"
    success_color: str = "92D050"
    border_color: str = "000000"


# Predefined themes
DEFAULT_THEME = ExcelTheme(
    header_fill="1B3A5C",
    header_font_color="FFFFFF",
    alt_row_fill="EBF2FA",
    accent_color="2E75B6",
    warning_color="FFC7CE",
    error_color="FF6B6B",
    success_color="92D050"
)

DARK_THEME = ExcelTheme(
    header_fill="1A1A1A",
    header_font_color="FFFFFF",
    alt_row_fill="2D2D2D",
    accent_color="4A90E2",
    warning_color="FFC7CE",
    error_color="FF6B6B",
    success_color="92D050"
)


def create_workbook(theme: Optional[ExcelTheme] = None) -> Workbook:
    """
    Create a new Excel workbook.
    
    Args:
        theme: ExcelTheme to apply (defaults to DEFAULT_THEME)
        
    Returns:
        Openpyxl Workbook instance
    """
    return Workbook()


def add_header_row(
    worksheet,
    headers: List[str],
    theme: Optional[ExcelTheme] = None
) -> None:
    """
    Add styled header row to worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        headers: List of header strings
        theme: ExcelTheme to use for styling
    """
    if theme is None:
        theme = DEFAULT_THEME
    
    # Create fill and font for header
    header_fill = PatternFill(start_color=theme.header_fill, end_color=theme.header_fill, fill_type="solid")
    header_font = Font(bold=True, color=theme.header_font_color, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Create border
    thin_border = Border(
        left=Side(style='thin', color=theme.border_color),
        right=Side(style='thin', color=theme.border_color),
        top=Side(style='thin', color=theme.border_color),
        bottom=Side(style='thin', color=theme.border_color)
    )
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def add_data_row(
    worksheet,
    row_data: List[Any],
    row_idx: int,
    theme: Optional[ExcelTheme] = None
) -> None:
    """
    Add data row with alternating row colors.
    
    Args:
        worksheet: Openpyxl worksheet object
        row_data: List of values for the row
        row_idx: Row index (1-based)
        theme: ExcelTheme to use for styling
    """
    if theme is None:
        theme = DEFAULT_THEME
    
    # Alternate row coloring (skip header, every other row)
    if (row_idx - 1) % 2 == 0:
        row_fill = PatternFill(start_color=theme.alt_row_fill, end_color=theme.alt_row_fill, fill_type="solid")
    else:
        row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Create border
    thin_border = Border(
        left=Side(style='thin', color=theme.border_color),
        right=Side(style='thin', color=theme.border_color),
        top=Side(style='thin', color=theme.border_color),
        bottom=Side(style='thin', color=theme.border_color)
    )
    
    row_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Add data cells
    for col_idx, value in enumerate(row_data, 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = row_alignment


def auto_column_widths(worksheet, max_width: int = 50) -> None:
    """
    Auto-fit column widths based on content.
    
    Args:
        worksheet: Openpyxl worksheet object
        max_width: Maximum column width (default 50)
    """
    for col_idx, column in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        
        # Set column width with some padding
        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def freeze_header(worksheet) -> None:
    """
    Freeze the header row in worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
    """
    worksheet.freeze_panes = "A2"


def add_severity_formatting(worksheet, severity_col: int) -> None:
    """
    Add conditional formatting based on severity column.
    
    Args:
        worksheet: Openpyxl worksheet object
        severity_col: Column index containing severity levels (1-based)
    """
    from openpyxl.formatting.rule import CellIsRule
    
    # Define severity colors
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Get the severity column letter
    col_letter = get_column_letter(severity_col)
    
    # Get the range for all data rows
    max_row = worksheet.max_row
    if max_row > 1:
        # Apply conditional formatting
        # Note: openpyxl's conditional formatting is limited, so we apply direct formatting
        for row_idx in range(2, max_row + 1):
            cell = worksheet.cell(row=row_idx, column=severity_col)
            if cell.value:
                value = str(cell.value).lower()
                if "error" in value or "critical" in value:
                    cell.fill = red_fill
                elif "warning" in value:
                    cell.fill = yellow_fill
                elif "info" in value or "low" in value:
                    cell.fill = green_fill


def save_workbook(workbook: Workbook, path: str) -> None:
    """
    Save workbook to file.
    
    Args:
        workbook: Openpyxl Workbook instance
        path: File path to save to
    """
    workbook.save(path)
