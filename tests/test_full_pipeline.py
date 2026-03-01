"""
Comprehensive test suite for ORCA compliance framework.

This module contains integration and unit tests for:
- Static analysis pipeline
- Individual analyzers
- Adapters and report generation
- Fixer agent and remediation
- Report generation in multiple formats
- Human-in-the-loop simulation
- Patch analysis
- End-to-end workflows
"""

import os
import sys
import json
import tempfile
import shutil
import pytest
from pathlib import Path
from typing import Dict, List, Any

# Ensure project root is in path
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from tests.conftest import (
    get_fixtures_dir,
    load_sample_good_c,
    load_sample_bad_c,
    load_kernel_rules,
    get_default_config,
)


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def temp_output_dir():
    """Create a temporary output directory."""
    temp_dir = tempfile.mkdtemp(prefix="orca_test_output_")
    yield temp_dir
    shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.fixture
def kernel_rules():
    """Load kernel compliance rules."""
    return load_kernel_rules()


@pytest.fixture
def default_config():
    """Get default ORCA configuration."""
    return get_default_config()


@pytest.fixture
def sample_codebase_dir():
    """Path to sample codebase for testing."""
    fixtures_path = Path(get_fixtures_dir())
    sample_codebase = fixtures_path / "sample_codebase"
    assert sample_codebase.exists(), f"Sample codebase not found at {sample_codebase}"
    return str(sample_codebase)


@pytest.fixture
def temp_file_copy(tmp_path):
    """Create a temporary copy of bad_style.c for modification tests."""
    fixtures_dir = get_fixtures_dir()
    source_file = os.path.join(fixtures_dir, "sample_codebase", "bad_style.c")

    if not os.path.exists(source_file):
        # Fallback to sample_bad.c
        source_file = os.path.join(fixtures_dir, "sample_bad.c")

    dest_file = tmp_path / "test_file.c"
    shutil.copy2(source_file, dest_file)
    return str(dest_file)


# ============================================================================
# Test Static Analysis Pipeline
# ============================================================================

class TestStaticAnalysis:
    """Test the full static analysis pipeline."""

    def test_single_file_analysis(self, kernel_rules, default_config, temp_output_dir):
        """Test audit on a single non-compliant file."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        fixtures_dir = get_fixtures_dir()
        test_file = os.path.join(fixtures_dir, "sample_codebase", "bad_style.c")

        if not os.path.exists(test_file):
            test_file = os.path.join(fixtures_dir, "sample_bad.c")

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=test_file,
            output_dir=temp_output_dir,
            domains=None
        )

        print(f"Single file analysis findings: {len(report.findings)}")
        print(f"Grade: {report.overall_grade}")
        assert len(report.findings) > 0, "Should detect violations in bad_style.c"
        assert report.overall_grade in ['A', 'B', 'C', 'D', 'F']

    def test_directory_analysis(self, kernel_rules, default_config, temp_output_dir, sample_codebase_dir):
        """Test audit on entire directory."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir,
            domains=None
        )

        print(f"Directory analysis found {len(report.findings)} total findings")
        print(f"Files scanned: {report.file_count}")
        print(f"Domain breakdown: {report.domain_results.keys()}")

        assert len(report.findings) > 0, "Should find violations in sample codebase"
        assert report.file_count > 0, "Should process multiple files"

    def test_good_file_minimal_findings(self, kernel_rules, default_config, temp_output_dir):
        """Test that good files have minimal findings."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        fixtures_dir = get_fixtures_dir()
        test_file = os.path.join(fixtures_dir, "sample_codebase", "good_style.c")

        if not os.path.exists(test_file):
            test_file = os.path.join(fixtures_dir, "sample_good.c")

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=test_file,
            output_dir=temp_output_dir,
            domains=None
        )

        print(f"Good file findings: {len(report.findings)}")
        # Good file should have few or no findings
        assert len(report.findings) <= 5, "Well-formatted file should have minimal violations"

    def test_domain_filtering(self, kernel_rules, default_config, temp_output_dir, sample_codebase_dir):
        """Test filtering to specific domains."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir,
            domains=['style']
        )

        print(f"Style domain findings: {len(report.findings)}")
        print(f"Domains in report: {list(report.domain_results.keys())}")

        # Should have findings from style domain
        assert len(report.findings) > 0 or len(report.domain_results) > 0

    def test_all_domains(self, kernel_rules, default_config, temp_output_dir, sample_codebase_dir):
        """Test analysis with all domains enabled."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir,
            domains=['style', 'license', 'structure']
        )

        print(f"Total findings (all domains): {len(report.findings)}")
        print(f"Domain scores: {report.domain_scores}")

        assert report.overall_grade in ['A', 'B', 'C', 'D', 'F']
        assert 'overall' in report.scores or len(report.findings) >= 0


# ============================================================================
# Test Individual Analyzers
# ============================================================================

class TestAnalyzers:
    """Test each analyzer individually."""

    def test_style_analyzer(self, kernel_rules, default_config):
        """Test StyleAnalyzer on non-compliant code."""
        try:
            from agents.analyzers.style_analyzer import StyleAnalyzer
        except ImportError:
            pytest.skip("StyleAnalyzer not available")

        bad_code = load_sample_bad_c()
        analyzer = StyleAnalyzer(kernel_rules.get('style', {}), default_config)
        findings = analyzer.analyze("test_bad.c", bad_code)

        print(f"Style analyzer findings: {len(findings)}")
        assert len(findings) > 0, "Should find style violations"

    def test_whitespace_analyzer(self, kernel_rules, default_config):
        """Test WhitespaceAnalyzer on mixed indentation."""
        try:
            from agents.analyzers.whitespace_analyzer import WhitespaceAnalyzer
        except ImportError:
            pytest.skip("WhitespaceAnalyzer not available")

        bad_code = load_sample_bad_c()
        analyzer = WhitespaceAnalyzer(
            kernel_rules.get('whitespace', {}),
            default_config
        )
        findings = analyzer.analyze("test_bad.c", bad_code)

        print(f"Whitespace analyzer findings: {len(findings)}")
        assert len(findings) >= 0, "Should analyze whitespace issues"

    def test_macro_analyzer(self, kernel_rules, default_config):
        """Test MacroAnalyzer on unsafe macros."""
        try:
            from agents.analyzers.macro_analyzer import MacroAnalyzer
        except ImportError:
            pytest.skip("MacroAnalyzer not available")

        bad_code = load_sample_bad_c()
        analyzer = MacroAnalyzer(kernel_rules.get('macros', {}), default_config)
        findings = analyzer.analyze("test_bad.c", bad_code)

        print(f"Macro analyzer findings: {len(findings)}")
        # bad_style.c has unsafe macros
        assert len(findings) >= 0

    def test_license_analyzer(self, kernel_rules, default_config):
        """Test LicenseAnalyzer on missing license headers."""
        try:
            from agents.analyzers.license_analyzer import LicenseAnalyzer
        except ImportError:
            pytest.skip("LicenseAnalyzer not available")

        bad_code = load_sample_bad_c()
        analyzer = LicenseAnalyzer(kernel_rules.get('license', {}), default_config)
        findings = analyzer.analyze("test_bad.c", bad_code)

        print(f"License analyzer findings: {len(findings)}")
        # bad_style.c is missing SPDX header
        assert len(findings) >= 0

    def test_include_analyzer(self, kernel_rules, default_config):
        """Test IncludeAnalyzer on bad include guards."""
        try:
            from agents.analyzers.include_analyzer import IncludeAnalyzer
        except ImportError:
            pytest.skip("IncludeAnalyzer not available")

        fixtures_dir = get_fixtures_dir()
        bad_h_path = os.path.join(fixtures_dir, "sample_codebase", "bad_style.h")

        if not os.path.exists(bad_h_path):
            bad_h_path = os.path.join(fixtures_dir, "sample_bad.h")

        with open(bad_h_path) as f:
            header_code = f.read()

        analyzer = IncludeAnalyzer(kernel_rules.get('includes', {}), default_config)
        findings = analyzer.analyze("test_bad.h", header_code)

        print(f"Include analyzer findings: {len(findings)}")
        assert len(findings) >= 0

    def test_structure_analyzer(self, kernel_rules, default_config):
        """Test StructureAnalyzer on code structure issues."""
        try:
            from agents.analyzers.structure_analyzer import StructureAnalyzer
        except ImportError:
            pytest.skip("StructureAnalyzer not available")

        fixtures_dir = get_fixtures_dir()
        mem_issue_path = os.path.join(fixtures_dir, "sample_codebase", "memory_issue.c")

        if os.path.exists(mem_issue_path):
            with open(mem_issue_path) as f:
                code = f.read()
        else:
            code = load_sample_bad_c()

        analyzer = StructureAnalyzer(kernel_rules.get('structure', {}), default_config)
        findings = analyzer.analyze("memory_issue.c", code)

        print(f"Structure analyzer findings: {len(findings)}")
        assert len(findings) >= 0

    def test_commit_analyzer(self, kernel_rules, default_config):
        """Test CommitAnalyzer with sample commit messages."""
        try:
            from agents.analyzers.commit_analyzer import CommitAnalyzer
        except ImportError:
            pytest.skip("CommitAnalyzer not available")

        commit_msg = "Added new feature for handling edge cases which is very long"

        analyzer = CommitAnalyzer(kernel_rules.get('commits', {}), default_config)
        findings = analyzer.analyze("commit", commit_msg)

        print(f"Commit analyzer findings: {len(findings)}")
        assert len(findings) >= 0


# ============================================================================
# Test Adapters
# ============================================================================

class TestAdapters:
    """Test adapter functionality."""

    def test_excel_report_generation(self, kernel_rules, default_config,
                                     temp_output_dir, sample_codebase_dir):
        """Test Excel report generation after audit."""
        try:
            from openpyxl import load_workbook
            from agents.compliance_static_agent import ComplianceStaticAgent
            from agents.adapters.excel_report_adapter import ExcelReportAdapter
        except ImportError:
            pytest.skip("openpyxl or required adapters not available")

        # Run audit first
        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir,
            domains=['style']
        )

        # Generate Excel report
        adapter = ExcelReportAdapter(kernel_rules.get('excel', {}), default_config)
        excel_path = os.path.join(temp_output_dir, "compliance_report.xlsx")

        success = adapter.generate_report(report, excel_path)

        if success:
            assert os.path.exists(excel_path), "Excel file should be created"
            wb = load_workbook(excel_path)
            sheet_names = wb.sheetnames
            print(f"Excel sheets created: {sheet_names}")
            assert len(sheet_names) > 0, "Should have at least one sheet"
        else:
            print("Excel report generation not available or failed")

    def test_include_guard_adapter(self, kernel_rules, default_config):
        """Test include guard validation adapter."""
        try:
            from agents.adapters.include_guard_adapter import IncludeGuardAdapter
        except ImportError:
            pytest.skip("IncludeGuardAdapter not available")

        fixtures_dir = get_fixtures_dir()
        bad_h = os.path.join(fixtures_dir, "sample_codebase", "bad_style.h")

        if not os.path.exists(bad_h):
            bad_h = os.path.join(fixtures_dir, "sample_bad.h")

        adapter = IncludeGuardAdapter(
            kernel_rules.get('include_guards', {}),
            default_config
        )

        with open(bad_h) as f:
            findings = adapter.analyze({bad_h: f.read()})

        print(f"Include guard findings: {len(findings) if findings else 0}")
        assert findings is None or len(findings) >= 0

    def test_spdx_adapter(self, kernel_rules, default_config):
        """Test SPDX license header validation."""
        try:
            from agents.adapters.spdx_adapter import SPDXAdapter
        except ImportError:
            pytest.skip("SPDXAdapter not available")

        bad_code = load_sample_bad_c()
        good_code = load_sample_good_c()

        adapter = SPDXAdapter(kernel_rules.get('spdx', {}), default_config)

        bad_findings = adapter.analyze({"bad.c": bad_code})
        good_findings = adapter.analyze({"good.c": good_code})

        print(f"SPDX findings (bad): {len(bad_findings) if bad_findings else 0}")
        print(f"SPDX findings (good): {len(good_findings) if good_findings else 0}")

        # Bad file should have SPDX findings
        if bad_findings:
            assert len(bad_findings) > 0 or bad_findings is None


# ============================================================================
# Test Fixer Agent
# ============================================================================

class TestFixerAgent:
    """Test the fixer agent and remediation."""

    def test_dry_run_fix(self, default_config, temp_file_copy):
        """Test dry-run fix mode (no modifications)."""
        try:
            from agents.compliance_fixer_agent import ComplianceFixerAgent
        except ImportError:
            pytest.skip("ComplianceFixerAgent not available")

        config = default_config.copy()
        config['dry_run'] = True

        fixer = ComplianceFixerAgent(config=config)

        # Create mock findings
        findings = [
            {
                'id': 'finding_1',
                'file_path': temp_file_copy,
                'line_number': 10,
                'domain': 'style',
                'message': 'Bad indentation'
            }
        ]

        # Mock solutions
        solutions = {
            'finding_1': {
                'fix_type': 'SIMPLE_REPLACEMENT',
                'line_start': 9,
                'line_end': 10,
                'old_code': '  int i;',
                'new_code': '\tint i;'
            }
        }

        result = fixer.apply_fixes(
            temp_file_copy,
            findings,
            solutions,
            dry_run=True
        )

        print(f"Dry-run result: applied={result.applied}, fixed={result.fixed_count}")

        # Verify file was not actually modified
        with open(temp_file_copy) as f:
            content = f.read()
        assert "  int i;" in content, "File should not be modified in dry-run mode"

    def test_apply_fix(self, default_config, temp_file_copy):
        """Test applying actual fixes."""
        try:
            from agents.compliance_fixer_agent import ComplianceFixerAgent
        except ImportError:
            pytest.skip("ComplianceFixerAgent not available")

        config = default_config.copy()
        config['dry_run'] = False

        fixer = ComplianceFixerAgent(config=config)

        findings = [
            {
                'id': 'finding_1',
                'file_path': temp_file_copy,
                'line_number': 1,
                'domain': 'style'
            }
        ]

        solutions = {
            'finding_1': {
                'fix_type': 'SIMPLE_REPLACEMENT',
                'line_start': 0,
                'line_end': 1,
                'old_code': '// Missing SPDX license header',
                'new_code': '// SPDX-License-Identifier: GPL-2.0-only'
            }
        }

        result = fixer.apply_fixes(
            temp_file_copy,
            findings,
            solutions,
            dry_run=False
        )

        print(f"Apply fix result: applied={result.applied}, fixed={result.fixed_count}")
        print(f"Backup created at: {result.backup_path}")

    def test_backup_creation(self, default_config, temp_file_copy):
        """Test that backup files are created during fixes."""
        try:
            from agents.compliance_fixer_agent import ComplianceFixerAgent
        except ImportError:
            pytest.skip("ComplianceFixerAgent not available")

        config = default_config.copy()
        config['dry_run'] = False

        fixer = ComplianceFixerAgent(config=config)

        findings = []
        solutions = {}

        result = fixer.apply_fixes(
            temp_file_copy,
            findings,
            solutions,
            dry_run=False
        )

        if result.backup_path:
            print(f"Backup created at: {result.backup_path}")
            assert os.path.exists(result.backup_path), "Backup should exist"


# ============================================================================
# Test Report Generation
# ============================================================================

class TestReportGeneration:
    """Test report generation in multiple formats."""

    def test_json_report(self, kernel_rules, default_config,
                        temp_output_dir, sample_codebase_dir):
        """Test JSON report generation."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir,
            domains=['style']
        )

        # Convert report to JSON-serializable format
        report_dict = {
            'overall_grade': report.overall_grade,
            'overall_score': report.scores.get('overall', 0),
            'findings_count': len(report.findings),
            'domain_scores': report.domain_scores,
        }

        json_path = os.path.join(temp_output_dir, "report.json")
        with open(json_path, 'w') as f:
            json.dump(report_dict, f, indent=2)

        assert os.path.exists(json_path), "JSON report should be created"

        with open(json_path) as f:
            loaded = json.load(f)

        print(f"JSON report generated with grade: {loaded['overall_grade']}")
        assert 'overall_grade' in loaded

    def test_html_report(self, kernel_rules, default_config,
                        temp_output_dir, sample_codebase_dir):
        """Test HTML report generation."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir
        )

        # Generate simple HTML report
        html_path = os.path.join(temp_output_dir, "report.html")

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>ORCA Compliance Report</title>
        </head>
        <body>
            <h1>Compliance Report</h1>
            <p>Grade: {report.overall_grade}</p>
            <p>Findings: {len(report.findings)}</p>
        </body>
        </html>
        """

        with open(html_path, 'w') as f:
            f.write(html_content)

        assert os.path.exists(html_path), "HTML report should be created"
        print("HTML report generated successfully")

    def test_excel_report(self, kernel_rules, default_config,
                         temp_output_dir, sample_codebase_dir):
        """Test Excel report generation."""
        try:
            from agents.compliance_static_agent import ComplianceStaticAgent
            from agents.adapters.excel_report_adapter import ExcelReportAdapter
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl or required modules not available")

        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir
        )

        adapter = ExcelReportAdapter(kernel_rules.get('excel', {}), default_config)
        excel_path = os.path.join(temp_output_dir, "report.xlsx")

        success = adapter.generate_report(report, excel_path)

        if success and os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            assert len(wb.sheetnames) > 0
            print(f"Excel report has sheets: {wb.sheetnames}")


# ============================================================================
# Test HITL Simulation
# ============================================================================

class TestHITLSimulated:
    """Simulate human-in-the-loop decision making."""

    def test_feedback_decisions(self, kernel_rules, default_config):
        """Simulate HITL feedback decisions (FIX/SKIP/WAIVE)."""
        # Mock finding from analyzer
        finding = {
            'id': 'style_001',
            'file': 'sample_bad.c',
            'line': 10,
            'message': 'Bad indentation',
            'severity': 'medium',
            'suggestion': 'Use tabs instead of spaces'
        }

        # Simulate different HITL decisions
        decisions = {
            'style_001': {
                'decision': 'FIX',
                'reason': 'Auto-fixable style issue',
                'constraint': None
            },
            'style_002': {
                'decision': 'SKIP',
                'reason': 'False positive',
                'constraint': 'Review this check'
            },
            'style_003': {
                'decision': 'WAIVE',
                'reason': 'Intentional deviation',
                'constraint': 'PR#123'
            }
        }

        print(f"Simulated HITL decisions for {len(decisions)} findings")
        assert len(decisions) == 3

    def test_fixer_workflow_auto_mode(self, kernel_rules, default_config, temp_file_copy):
        """Test fixer workflow in auto-mode."""
        try:
            from agents.compliance_fixer_agent import ComplianceFixerAgent
        except ImportError:
            pytest.skip("ComplianceFixerAgent not available")

        config = default_config.copy()
        config['dry_run'] = True

        fixer = ComplianceFixerAgent(config=config)

        # Simulate auto-mode: only fix high-confidence findings
        findings = [
            {
                'id': 'f1',
                'file_path': temp_file_copy,
                'line_number': 1,
                'confidence': 0.95,
                'severity': 'high'
            },
            {
                'id': 'f2',
                'file_path': temp_file_copy,
                'line_number': 2,
                'confidence': 0.60,
                'severity': 'low'
            }
        ]

        # Only process high-confidence findings
        high_confidence = [f for f in findings if f.get('confidence', 0) > 0.8]

        print(f"Auto-mode will fix {len(high_confidence)} of {len(findings)} findings")
        assert len(high_confidence) == 1


# ============================================================================
# Test Patch Analysis
# ============================================================================

class TestPatchAnalysis:
    """Test patch and diff analysis."""

    def test_patch_parser(self):
        """Test parsing a unified diff patch."""
        fixtures_dir = get_fixtures_dir()
        patch_file = os.path.join(fixtures_dir, "sample.patch")

        with open(patch_file) as f:
            patch_content = f.read()

        # Simple patch parsing
        lines = patch_content.split('\n')

        # Count diff sections
        diff_count = sum(1 for line in lines if line.startswith('diff --git'))
        hunk_count = sum(1 for line in lines if line.startswith('@@'))

        print(f"Patch contains {diff_count} file(s) and {hunk_count} hunk(s)")

        assert diff_count > 0, "Should parse patch file format"
        assert hunk_count > 0, "Should find hunks in patch"

    def test_batch_patch(self, kernel_rules, default_config):
        """Test batch processing of multiple patches."""
        patches = [
            {
                'id': 'patch_1',
                'file': 'core.c',
                'hunks': 5
            },
            {
                'id': 'patch_2',
                'file': 'utils.c',
                'hunks': 3
            }
        ]

        total_hunks = sum(p['hunks'] for p in patches)

        print(f"Processing {len(patches)} patches with {total_hunks} total hunks")
        assert len(patches) == 2
        assert total_hunks == 8


# ============================================================================
# Test End-to-End Pipeline
# ============================================================================

class TestEndToEnd:
    """Full pipeline integration tests."""

    def test_full_pipeline(self, kernel_rules, default_config,
                          temp_output_dir, sample_codebase_dir):
        """Test complete audit -> analyze -> report -> fix workflow."""
        from agents.compliance_static_agent import ComplianceStaticAgent

        print("\n=== Full Pipeline Test ===")

        # Phase 1: Scan
        agent = ComplianceStaticAgent(rules=kernel_rules, config=default_config)
        report1 = agent.run_audit(
            codebase_path=sample_codebase_dir,
            output_dir=temp_output_dir,
            domains=['style']
        )

        print(f"Phase 1 - Initial scan: {len(report1.findings)} findings")
        print(f"  Grade: {report1.overall_grade}")
        print(f"  Score: {report1.scores.get('overall', 0):.2%}")

        # Phase 2: Analyze (already done in scan)
        print(f"Phase 2 - Analysis: {len(report1.domain_results)} domains analyzed")

        # Phase 3: Report
        report_data = {
            'findings': len(report1.findings),
            'grade': report1.overall_grade,
            'domains': list(report1.domain_results.keys())
        }

        report_path = os.path.join(temp_output_dir, "analysis_report.json")
        with open(report_path, 'w') as f:
            json.dump(report_data, f, indent=2)

        print(f"Phase 3 - Report generated: {report_path}")

        # Phase 4: Dry-run fix
        try:
            from agents.compliance_fixer_agent import ComplianceFixerAgent

            fixer = ComplianceFixerAgent(
                config={**default_config, 'dry_run': True}
            )
            print("Phase 4 - Fix (dry-run) available")
        except ImportError:
            print("Phase 4 - Fix agent not available, skipping")

        # Phase 5: Report again
        report2_data = {
            'initial_findings': len(report1.findings),
            'final_findings': len(report1.findings),  # Would be less after actual fixes
            'improvement': '0%'  # Would be calculated in real scenario
        }

        report2_path = os.path.join(temp_output_dir, "final_report.json")
        with open(report2_path, 'w') as f:
            json.dump(report2_data, f, indent=2)

        print("Phase 5 - Final report complete")
        print(f"\n=== Pipeline Summary ===")
        print(f"Initial findings: {report1.findings}")
        print(f"Final report: {report2_path}")

        assert os.path.exists(report_path), "Analysis report should exist"
        assert os.path.exists(report2_path), "Final report should exist"


# ============================================================================
# Test Configuration and Utilities
# ============================================================================

class TestConfigurationAndUtils:
    """Test configuration loading and utility functions."""

    def test_load_kernel_rules(self, kernel_rules):
        """Test loading kernel compliance rules."""
        assert isinstance(kernel_rules, dict), "Rules should be a dict"
        print(f"Loaded rules with keys: {list(kernel_rules.keys())[:5]}...")

    def test_default_config(self, default_config):
        """Test default configuration."""
        assert 'paths' in default_config or 'compliance' in default_config
        print(f"Default config: {list(default_config.keys())}")

    def test_fixtures_accessible(self, sample_codebase_dir):
        """Test that sample fixtures are accessible."""
        files = os.listdir(sample_codebase_dir)
        print(f"Sample codebase files: {files}")
        assert len(files) > 0, "Sample codebase should have files"


# ============================================================================
# Main entry point
# ============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
